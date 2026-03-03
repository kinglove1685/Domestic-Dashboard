from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
import re
import html
from typing import Optional

import altair as alt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="국내영업 생산요청 vs 출고 관리 TEST", layout="wide")
st.title("국내영업 생산요청 대비 출고 관리 TEST")
st.caption("기준: 품목코드 앞 4자리(예: S129) 동일 시 1 EA = 1 PACK, 분기 누적 출고 집계")
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');
    :root {
        --font: "Noto Sans KR", sans-serif;
        --bg1: #f8fbff;
        --bg2: #eef5ff;
        --card: #ffffff;
        --border: #dbe7ff;
        --accent: #1d4ed8;
    }
    html, body, .stApp, [class*="st-"], [class*="css"] {
        font-family: "Noto Sans KR", sans-serif !important;
    }
    .stApp {
        background: linear-gradient(180deg, var(--bg1) 0%, var(--bg2) 100%);
    }
    div[data-testid="metric-container"] {
        background: var(--card);
        border: 1px solid var(--border);
        border-radius: 12px;
        padding: 10px 14px;
        box-shadow: 0 2px 8px rgba(29, 78, 216, 0.06);
    }
    div[data-testid="metric-container"] label {
        color: #475569 !important;
        font-weight: 600 !important;
    }
    div[data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #0f172a;
    }
    button[data-baseweb="tab"] {
        font-weight: 700;
        color: #334155;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: var(--accent);
        background: #eaf2ff;
        border-radius: 8px;
    }
    div[data-testid="stTextInput"] input {
        border: 1px solid #bfdbfe;
        background: #f8fbff;
    }
    div[data-testid="stTextInput"] input:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 1px var(--accent);
    }
    </style>
    """,
    unsafe_allow_html=True,
)

APP_FONT_FAMILY = "Noto Sans KR"


def normalize_name(name: object) -> str:
    return str(name).replace("\n", "").replace(" ", "").replace("\t", "").strip()


def build_colmap(df: pd.DataFrame) -> dict[str, str]:
    return {normalize_name(c): c for c in df.columns}


def find_col(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    cmap = build_colmap(df)
    for c in candidates:
        key = normalize_name(c)
        if key in cmap:
            return cmap[key]
    return None


def to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)


def normalize_code(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip()


PACK_COUNT_PATTERNS = [
    re.compile(r"(?<!\d)(\d{1,3})\s*팩"),
    re.compile(r"(?<!\d)(\d{1,3})\s*개입"),
    re.compile(r"(?<!\d)(\d{1,3})\s*P\b", re.IGNORECASE),
    re.compile(r"_(\d{1,3})(?:\b|$)"),
]
PACK_COUNT_PCODE_OVERRIDES = {
    "P1136": 1.0,
}


def extract_pack_count(name: object) -> float:
    if pd.isna(name):
        return 1.0
    text = str(name).strip()
    if not text:
        return 1.0
    for pattern in PACK_COUNT_PATTERNS:
        match = pattern.search(text)
        if match:
            value = int(match.group(1))
            if value > 0:
                return float(value)
    return 1.0


def apply_pack_count_overrides_by_pcode(df: pd.DataFrame, pcode_col: str = "P코드", pack_col: str = "PACK당낱개수") -> pd.DataFrame:
    out = df.copy()
    if pcode_col not in out.columns or pack_col not in out.columns:
        return out
    pcode_series = out[pcode_col].fillna("").astype(str)
    for pcode, forced_pack in PACK_COUNT_PCODE_OVERRIDES.items():
        mask = pcode_series.str.contains(rf"(^|,\s*){re.escape(pcode)}(\s*,|$)", regex=True)
        out.loc[mask, pack_col] = float(forced_pack)
    return out


def normalize_product_family_name(name: object) -> str:
    if pd.isna(name):
        return ""
    text = str(name).strip()
    if not text:
        return ""
    # leading quantity token (e.g. 10팩_Iris ...)
    text = re.sub(r"(?i)^\s*\d{1,3}\s*(팩|P|개입)\s*[_\-\s]*", "", text)
    # trailing quantity token (e.g. ..._40팩, ..._30P, ..._30개입)
    text = re.sub(r"(?i)[_\-\s]*\d{1,3}\s*(팩|P|개입)\s*$", "", text)
    # trailing bare numeric token (e.g. ..._30)
    text = re.sub(r"(?i)[_\-\s]*\d{1,3}\s*$", "", text)
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def summarize_pack_counts(values: pd.Series, max_units: int = 4) -> str:
    nums = pd.to_numeric(values, errors="coerce").dropna()
    units = sorted({int(v) for v in nums if int(v) > 0})
    if not units:
        return ""
    if len(units) <= max_units:
        return ", ".join(str(v) for v in units)
    return ", ".join(str(v) for v in units[:max_units]) + f" 외 {len(units) - max_units}"


def summarize_names(values: pd.Series, max_names: int = 3) -> str:
    names = [str(v).strip() for v in values if pd.notna(v) and str(v).strip() and str(v).strip() != "0"]
    if not names:
        return ""
    seen = []
    for n in names:
        if n not in seen:
            seen.append(n)
    if len(seen) <= max_names:
        return ", ".join(seen)
    return ", ".join(seen[:max_names]) + f" 외 {len(seen) - max_names}"


def summarize_codes(values: pd.Series, max_codes: int = 5) -> str:
    codes = [str(v).strip() for v in values if pd.notna(v) and str(v).strip() and str(v).strip() != "0"]
    if not codes:
        return ""
    seen = []
    for c in codes:
        if c not in seen:
            seen.append(c)
    if len(seen) <= max_codes:
        return ", ".join(seen)
    return ", ".join(seen[:max_codes]) + f" 외 {len(seen) - max_codes}"


def merge_text_values(*values: object) -> str:
    out = []
    for v in values:
        s = str(v).strip()
        if not s or s == "0":
            continue
        if s not in out:
            out.append(s)
    return ", ".join(out)


def normalize_key_value(value: object) -> str:
    s = str(value).strip()
    if s.lower() in {"", "0", "nan", "none"}:
        return ""
    return s


def normalize_product_type(value: object) -> str:
    text = normalize_key_value(value)
    if not text:
        return "미지정"
    lowered = text.lower()
    if text == "본품" or lowered in {"정규품", "정규"}:
        return "정규품"
    if text == "샘플" or lowered in {"sample"}:
        return "샘플"
    return text


def split_master_codes(value: object) -> list[str]:
    text = normalize_key_value(value)
    if not text:
        return []
    parts = [p.strip() for p in str(text).split(",") if p and p.strip()]
    codes = []
    for part in parts:
        # summarize_codes 결과의 꼬리("... 외 27")를 제거
        part = re.sub(r"\s*외\s*\d+\s*$", "", part).strip()
        if not part:
            continue
        if part.startswith("외"):
            continue
        if part not in codes:
            codes.append(part)
    return codes


def calc_stock_sum_for_master_codes(master_code_value: object, stock_map: dict[str, float]) -> float:
    return float(sum(stock_map.get(code, 0.0) for code in split_master_codes(master_code_value)))


def build_pcode_stock_totals(item_df: pd.DataFrame, stock_map: dict[str, float]) -> dict[str, float]:
    if item_df is None or item_df.empty or not stock_map:
        return {}
    if "P코드" not in item_df.columns or "제품코드(마스터)" not in item_df.columns:
        return {}

    pcode_to_master_codes: dict[str, set[str]] = {}
    for _, row in item_df[["P코드", "제품코드(마스터)"]].iterrows():
        pcode_text = normalize_key_value(row["P코드"])
        if not pcode_text:
            continue
        pcode_list = split_master_codes(pcode_text)
        master_codes = split_master_codes(row["제품코드(마스터)"])
        if not pcode_list or not master_codes:
            continue
        for pcode in pcode_list:
            if pcode not in pcode_to_master_codes:
                pcode_to_master_codes[pcode] = set()
            pcode_to_master_codes[pcode].update(master_codes)

    return {
        pcode: float(sum(stock_map.get(code, 0.0) for code in master_codes))
        for pcode, master_codes in pcode_to_master_codes.items()
    }


def build_pcode_detail_map(family_df: pd.DataFrame) -> dict[str, str]:
    if family_df is None or family_df.empty:
        return {}
    if "P코드" not in family_df.columns or "집계키" not in family_df.columns:
        return {}

    pcode_to_details: dict[str, list[str]] = {}
    for _, row in family_df[["P코드", "집계키"]].iterrows():
        detail = normalize_key_value(row["집계키"])
        if not detail:
            continue
        for pcode in split_master_codes(row["P코드"]):
            if not pcode:
                continue
            if pcode not in pcode_to_details:
                pcode_to_details[pcode] = []
            if detail not in pcode_to_details[pcode]:
                pcode_to_details[pcode].append(detail)

    return {pcode: summarize_codes(pd.Series(details), max_codes=8) for pcode, details in pcode_to_details.items()}


@st.cache_data
def load_data(base_dir: str) -> tuple[pd.DataFrame, pd.DataFrame, str, str]:
    files = [f for f in Path(base_dir).glob("*.xlsx") if not f.name.startswith("~$")]
    if not files:
        raise FileNotFoundError("현재 폴더에 xlsx 파일이 없습니다.")

    request_df = None
    inbound_df = None
    request_file = ""
    inbound_file = ""

    for file in files:
        try:
            df = pd.read_excel(file)
        except Exception:
            continue
        cmap = build_colmap(df)
        has_item = normalize_name("품목코드") in cmap
        has_qty = normalize_name("수량") in cmap
        has_request_qty = any("요청수량" in normalize_name(c) for c in df.columns)

        if has_item and has_request_qty and not has_qty:
            request_df = df
            request_file = file.name
        elif has_item and has_qty:
            inbound_df = df
            inbound_file = file.name

    if request_df is None or inbound_df is None:
        raise ValueError("요청 파일/입고 파일을 자동 식별하지 못했습니다. xlsx 파일 구조를 확인해주세요.")

    return request_df, inbound_df, request_file, inbound_file


@st.cache_data
def load_item_product_master_map(base_dir: str) -> pd.DataFrame:
    candidates = list(Path(base_dir).glob("*마스터 데이터*.xlsx"))
    if not candidates:
        return pd.DataFrame(columns=["품목코드", "제품코드(마스터)"])

    for file in candidates:
        try:
            df = pd.read_excel(file)
        except Exception:
            continue

        item_col = find_col(df, ["품목코드"])
        product_col = find_col(df, ["제품코드"])
        if not item_col or not product_col:
            continue

        mapping = df[[item_col, product_col]].copy()
        mapping.columns = ["품목코드", "제품코드(마스터)"]
        mapping["품목코드"] = normalize_code(mapping["품목코드"])
        mapping["제품코드(마스터)"] = mapping["제품코드(마스터)"].fillna("").astype(str).str.strip()
        mapping = mapping[(mapping["품목코드"] != "")]
        mapping = (
            mapping.groupby("품목코드", as_index=False)["제품코드(마스터)"]
            .apply(summarize_codes)
            .rename(columns={"제품코드(마스터)": "제품코드(마스터)"})
        )
        mapping["제품코드(마스터)"] = mapping["제품코드(마스터)"].fillna("")
        return mapping

    return pd.DataFrame(columns=["품목코드", "제품코드(마스터)"])


@st.cache_data
def load_inventory_stock(base_dir: str) -> tuple[pd.DataFrame, str]:
    candidates = list(Path(base_dir).glob("*재고장*.xlsx"))
    if not candidates:
        return pd.DataFrame(columns=["제품코드(마스터)", "보유재고"]), ""

    for file in candidates:
        try:
            df = pd.read_excel(file)
        except Exception:
            continue

        code_col = find_col(df, ["품목코드"])
        stock_col = find_col(df, ["재고", "재고수량", "현재고"])
        if not code_col or not stock_col:
            continue

        stock = df[[code_col, stock_col]].copy()
        stock.columns = ["제품코드(마스터)", "보유재고"]
        stock["제품코드(마스터)"] = normalize_code(stock["제품코드(마스터)"])
        stock["보유재고"] = to_numeric(stock["보유재고"])
        stock = stock[stock["제품코드(마스터)"] != ""]
        stock = stock.groupby("제품코드(마스터)", as_index=False)["보유재고"].sum()
        return stock, file.name

    return pd.DataFrame(columns=["제품코드(마스터)", "보유재고"]), ""


def prepare_request(df: pd.DataFrame) -> pd.DataFrame:
    year_col = find_col(df, ["년"])
    quarter_col = find_col(df, ["분기"])
    item_col = find_col(df, ["품목코드"])
    name_col = find_col(df, ["품명"])
    pcode_col = find_col(df, ["P 코드", "P코드"])
    brand_col = find_col(df, ["브랜드", "BRAND", "Brand"])
    type_col = find_col(df, ["구분", "구 분"])

    if not (year_col and quarter_col and item_col):
        raise ValueError("요청 파일에 필수 컬럼(년/분기/품목코드)이 없습니다.")

    qty_cols = [c for c in df.columns if "요청수량" in normalize_name(c)]
    if not qty_cols:
        raise ValueError("요청 파일에서 요청수량 컬럼을 찾지 못했습니다.")

    quarter_total_cols = [c for c in qty_cols if "분기" in normalize_name(c)]

    req = df.copy()
    req["년"] = pd.to_numeric(req[year_col], errors="coerce").astype("Int64")
    req["분기"] = pd.to_numeric(req[quarter_col], errors="coerce").astype("Int64")
    req["품목코드"] = normalize_code(req[item_col])
    req["제품코드"] = req["품목코드"].str[:4]
    req["P코드"] = req[pcode_col] if pcode_col else ""
    req["브랜드"] = req[brand_col] if brand_col else ""
    req["구분"] = req[type_col].apply(normalize_product_type) if type_col else "미지정"
    req["품명"] = req[name_col] if name_col else ""
    req["제품군명"] = req["품명"].apply(normalize_product_family_name)
    req["PACK당낱개수"] = req["품명"].apply(extract_pack_count)
    req = apply_pack_count_overrides_by_pcode(req, pcode_col="P코드", pack_col="PACK당낱개수")

    if quarter_total_cols:
        req["요청수량_PACK"] = to_numeric(req[quarter_total_cols[0]])
    else:
        qty_numeric = req[qty_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
        req["요청수량_PACK"] = qty_numeric.sum(axis=1)

    return req[["년", "분기", "품목코드", "제품코드", "P코드", "브랜드", "구분", "품명", "제품군명", "PACK당낱개수", "요청수량_PACK"]]


def prepare_inbound(df: pd.DataFrame) -> pd.DataFrame:
    year_col = find_col(df, ["년"])
    quarter_col = find_col(df, ["분기"])
    item_col = find_col(df, ["품목코드"])
    qty_col = find_col(df, ["수량"])
    date_col = find_col(df, ["이동일자"])
    name_col = find_col(df, ["품명"])

    if not (year_col and quarter_col and item_col and qty_col):
        raise ValueError("입고 파일에 필수 컬럼(년/분기/품목코드/수량)이 없습니다.")

    inbound = df.copy()
    inbound["년"] = pd.to_numeric(inbound[year_col], errors="coerce").astype("Int64")
    inbound["분기"] = pd.to_numeric(inbound[quarter_col], errors="coerce").astype("Int64")
    inbound["품목코드"] = normalize_code(inbound[item_col])
    inbound["제품코드"] = inbound["품목코드"].str[:4]
    inbound["브랜드"] = ""
    inbound["구분"] = "미지정"
    pcode_col = find_col(df, ["P 코드", "P코드"])
    inbound["P코드"] = inbound[pcode_col] if pcode_col else ""
    inbound["품명"] = inbound[name_col] if name_col else ""
    inbound["제품군명"] = inbound["품명"].apply(normalize_product_family_name)
    inbound["PACK당낱개수"] = inbound["품명"].apply(extract_pack_count)
    inbound = apply_pack_count_overrides_by_pcode(inbound, pcode_col="P코드", pack_col="PACK당낱개수")
    inbound["출고수량_EA"] = to_numeric(inbound[qty_col])
    inbound["이동일자"] = pd.to_datetime(inbound[date_col], errors="coerce") if date_col else pd.NaT

    return inbound[["년", "분기", "품목코드", "제품코드", "브랜드", "구분", "품명", "제품군명", "PACK당낱개수", "이동일자", "출고수량_EA"]]


def status_label(request_qty: pd.Series, shipped_qty: pd.Series) -> pd.Series:
    return np.select(
        [
            (request_qty == 0) & (shipped_qty > 0),
            (request_qty > 0) & (shipped_qty == 0),
            (request_qty > 0) & (shipped_qty < request_qty),
            (request_qty > 0) & (shipped_qty == request_qty),
            (request_qty > 0) & (shipped_qty > request_qty),
        ],
        [
            "요청없음(출고발생)",
            "미출고",
            "출고중",
            "출고완료",
            "요청초과출고",
        ],
        default="확인필요",
    )


def add_progress_columns(df: pd.DataFrame, req_col: str, ship_col: str) -> pd.DataFrame:
    out = df.copy()
    req = pd.to_numeric(out[req_col], errors="coerce").fillna(0)
    ship_total = pd.to_numeric(out[ship_col], errors="coerce").fillna(0)
    ship_matched = np.minimum(req, ship_total)
    ship_excess = np.maximum(ship_total - req, 0)

    out["총출고수량_EA"] = ship_total
    out["매칭출고수량_EA"] = ship_matched
    out["초과출고수량_EA"] = ship_excess
    out["잔량"] = np.maximum(req - ship_matched, 0)
    out["진행률(%)"] = np.where(req > 0, (ship_matched / req) * 100, np.nan)
    out["상태"] = status_label(req, ship_total)
    return out


def format_table(
    df: pd.DataFrame,
    int_cols: list[str],
    pct_cols: list[str] | None = None,
    progress_bar_cols: list[str] | None = None,
    status_col: str | None = None,
    positive_alert_cols: list[str] | None = None,
):
    pct_cols = pct_cols or []
    progress_bar_cols = progress_bar_cols or []
    positive_alert_cols = positive_alert_cols or []
    fmt = {c: "{:,.0f}" for c in int_cols if c in df.columns}
    fmt.update({c: "{:,.1f}" for c in pct_cols if c in df.columns})
    styler = df.style.format(fmt, na_rep="")
    for c in progress_bar_cols:
        if c in df.columns:
            # 100%를 가득 찬 기준으로 시각화(초과값은 막대가 가득 찬 상태로 표시)
            styler = styler.bar(subset=[c], vmin=0, vmax=100, color="#93c5fd")

    if status_col and status_col in df.columns:
        status_style_map = {
            "미출고": "background-color:#fee2e2;color:#991b1b;font-weight:700;",
            "출고중": "background-color:#fef3c7;color:#92400e;font-weight:700;",
            "출고완료": "background-color:#dcfce7;color:#166534;font-weight:700;",
            "요청초과출고": "background-color:#ffedd5;color:#9a3412;font-weight:700;",
            "요청없음(출고발생)": "background-color:#e2e8f0;color:#334155;font-weight:700;",
            "확인필요": "background-color:#f1f5f9;color:#334155;font-weight:700;",
        }
        styler = styler.applymap(lambda v: status_style_map.get(str(v), ""), subset=[status_col])

    def style_positive(v: object) -> str:
        num = pd.to_numeric(pd.Series([v]), errors="coerce").iloc[0]
        if pd.isna(num):
            return ""
        if float(num) > 0:
            return "color:#b91c1c;font-weight:700;"
        return ""

    for c in positive_alert_cols:
        if c in df.columns:
            styler = styler.applymap(style_positive, subset=[c])

    return styler


def parse_search_terms(query: str) -> list[str]:
    # 공백은 제품명 내부 문자로 취급하고, 구분자는 쉼표/|/ 로만 사용한다.
    terms = [t.strip() for t in re.split(r"[,|/]+", str(query)) if t.strip()]
    # keep input order while removing duplicates
    return list(dict.fromkeys(terms))


def apply_or_search(df: pd.DataFrame, query: str, columns: list[str]) -> pd.DataFrame:
    terms = parse_search_terms(query)
    use_cols = [c for c in columns if c in df.columns]
    if not terms or not use_cols:
        return df

    mask = pd.Series(False, index=df.index)
    for term in terms:
        term_mask = pd.Series(False, index=df.index)
        for col in use_cols:
            term_mask = term_mask | df[col].astype(str).str.contains(term, case=False, na=False, regex=False)
        mask = mask | term_mask
    return df[mask]


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "data", merge_cols: Optional[list[str]] = None) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # 숫자 컬럼은 엑셀에서 천단위 구분기호 포맷으로 저장
        for col_idx, col_name in enumerate(df.columns, start=1):
            series = df[col_name]
            if not pd.api.types.is_numeric_dtype(series):
                continue

            num_fmt = "#,##0.0" if "(%)" in str(col_name) else "#,##0"
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                cell.number_format = num_fmt

        # 계층형 병합(예: [생산, 판매, 개입, 품 명])이 필요한 경우 세로 병합 적용
        if merge_cols:
            valid_merge_cols = [c for c in merge_cols if c in df.columns]
            col_idx_map = {col_name: idx for idx, col_name in enumerate(df.columns, start=1)}

            def merge_value(v: object) -> str:
                if pd.isna(v):
                    return ""
                return str(v)

            for m_i, m_col in enumerate(valid_merge_cols):
                col_idx = col_idx_map[m_col]
                prev_cols = valid_merge_cols[:m_i]
                start = 0
                while start < len(df):
                    end = start + 1
                    while end < len(df):
                        same_curr = merge_value(df.iloc[end][m_col]) == merge_value(df.iloc[start][m_col])
                        same_prev = all(
                            merge_value(df.iloc[end][pc]) == merge_value(df.iloc[start][pc]) for pc in prev_cols
                        )
                        if not (same_curr and same_prev):
                            break
                        end += 1
                    if end - start > 1:
                        ws.merge_cells(
                            start_row=start + 2,  # row 1 = header
                            end_row=end + 1,
                            start_column=col_idx,
                            end_column=col_idx,
                        )
                    start = end
    buffer.seek(0)
    return buffer.getvalue()


def to_master_pack_plan_excel_bytes(
    prod_df: pd.DataFrame,
    pcode_stock_totals: Optional[dict[str, float]] = None,
    pcode_detail_map: Optional[dict[str, str]] = None,
    raw_req_df: Optional[pd.DataFrame] = None,
    detail_item_df: Optional[pd.DataFrame] = None,
    detail_stock_by_code: Optional[dict[str, float]] = None,
    raw_inbound_df: Optional[pd.DataFrame] = None,
) -> bytes:
    required_cols = [
        "제품코드",
        "P코드",
        "PACK당낱개수",
        "품명",
        "요청수량_PACK",
        "요청수량_낱개",
        "매칭출고수량_EA",
        "출고수량_낱개",
        "잔량",
        "잔량_낱개",
    ]
    src = prod_df.copy()
    for col in required_cols:
        if col not in src.columns:
            src[col] = ""
    code_col = required_cols[0]
    pcode_col = required_cols[1]
    name_col = required_cols[3]

    src["P코드"] = src["P코드"].fillna("").astype(str).str.strip()
    src["제품코드"] = src["제품코드"].fillna("").astype(str).str.strip()
    src["품명"] = src["품명"].fillna("").astype(str).str.strip()
    src["PACK당낱개수"] = src["PACK당낱개수"].fillna("").astype(str).str.strip()

    # 요약 문자열(예: "외27")로 축약된 P코드를 원본 요청 데이터에서 복원
    product_to_full_pcodes: dict[str, list[str]] = {}
    prefix_to_detail_codes: dict[str, list[str]] = {}
    if raw_req_df is not None and not raw_req_df.empty:
        req_pcode_col = find_col(raw_req_df, ["P코드", "P 코드"])
        req_prod_col = find_col(raw_req_df, ["제품코드", "품목코드"])
        req_item_col = find_col(raw_req_df, ["품목코드"])
        if req_pcode_col and req_prod_col:
            for _, rr in raw_req_df.iterrows():
                prod_code = normalize_key_value(rr.get(req_prod_col, ""))
                if req_item_col and req_prod_col == req_item_col and len(prod_code) >= 4:
                    prod_code = prod_code[:4]
                if not prod_code:
                    continue
                codes = split_master_codes(rr.get(req_pcode_col, ""))
                if not codes:
                    continue
                if prod_code not in product_to_full_pcodes:
                    product_to_full_pcodes[prod_code] = []
                for c in codes:
                    if c not in product_to_full_pcodes[prod_code]:
                        product_to_full_pcodes[prod_code].append(c)
    if detail_item_df is not None and not detail_item_df.empty:
        item_detail_col_for_prefix = find_col(detail_item_df, ["제품코드(마스터)", "세부코드"])
        if item_detail_col_for_prefix:
            for v in detail_item_df[item_detail_col_for_prefix]:
                code = normalize_key_value(v)
                if not code:
                    continue
                prefix = code.split("-")[0].strip()
                if not prefix:
                    continue
                if prefix not in prefix_to_detail_codes:
                    prefix_to_detail_codes[prefix] = []
                if code not in prefix_to_detail_codes[prefix]:
                    prefix_to_detail_codes[prefix].append(code)

    def expanded_master_codes(raw_pcode_text: object, rep_code: object) -> list[str]:
        base = split_master_codes(raw_pcode_text)
        rep = normalize_key_value(rep_code)
        extras = product_to_full_pcodes.get(rep, [])
        # raw 값이 "코드 외 N"인데 제품코드 매핑으로 못 찾은 경우, prefix 기반 fallback
        raw_text = normalize_key_value(raw_pcode_text)
        has_tail_count = bool(re.search(r"\s*외\s*\d+\s*$", raw_text))
        if has_tail_count and not extras and base:
            seed_prefix = str(base[0]).split("-")[0].strip()
            extras = prefix_to_detail_codes.get(seed_prefix, [])
        merged: list[str] = []
        for c in base + extras:
            if c and c not in merged:
                merged.append(c)
        return merged
    def pack_sort_value(value: object) -> float:
        text = str(value).strip()
        if not text:
            return -1.0
        nums = [int(n) for n in re.findall(r"\d+", text)]
        if not nums:
            return -1.0
        return float(max(nums))

    # P코드 묶음 개수가 많은 그룹을 우선 배치하고, 그룹 내에서는 개입(C) 높은 순으로 정렬
    src["_group_key"] = np.where(src["P코드"] != "", "P:" + src["P코드"], "ROW:" + src.index.astype(str))
    src["_group_size"] = src.groupby("_group_key", dropna=False)["_group_key"].transform("size")
    src["_pcode_sort"] = np.where(src["P코드"] != "", src["P코드"], "ZZZZZZZZ")
    src["_pack_sort"] = src["PACK당낱개수"].apply(pack_sort_value)
    src["_group_pack_max"] = src.groupby("_group_key", dropna=False)["_pack_sort"].transform("max")
    src = src.sort_values(
        ["_group_size", "_group_pack_max", "_pcode_sort", "P코드", "_pack_sort", "제품코드", "품명"],
        ascending=[False, False, True, True, False, True, True],
    ).reset_index(drop=True)
    pcode_stock_totals = pcode_stock_totals or {}

    def to_int_or_none(value: object) -> Optional[int]:
        n = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        if pd.isna(n):
            return None
        return int(round(float(n)))

    group_sum_j = (
        src.groupby("_group_key", dropna=False)["잔량_낱개"]
        .sum(min_count=1)
        .apply(lambda v: 0 if pd.isna(v) else int(round(float(v))))
        .to_dict()
    )
    group_stock_l_raw = {
        gk: float(sum(pcode_stock_totals.get(code, 0.0) for code in split_master_codes(str(gk)[2:])))
        if str(gk).startswith("P:")
        else 0.0
        for gk in group_sum_j.keys()
    }
    # 엑셀 내보내기에서는 필요 수량(K)을 초과한 재고는 표시하지 않음
    group_stock_l = {
        gk: float(min(float(group_stock_l_raw.get(gk, 0.0)), float(group_sum_j.get(gk, 0))))
        for gk in group_sum_j.keys()
    }
    # 생산 필요량(pcs) = 오더 낱개 총량(K) - 낱개 재고(본사)(L), 양수는 유지, 음수면 0
    group_need_m = {
        gk: int(round(max(float(group_sum_j.get(gk, 0)) - float(group_stock_l.get(gk, 0.0)), 0.0)))
        for gk in group_sum_j.keys()
    }
    src["_pcode_stock_total"] = src["_group_key"].apply(lambda gk: float(group_stock_l.get(gk, 0.0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Master 포장계획 요약"

    header_row = 3
    data_start_row = 4
    master_headers = {
        1: "판매",
        2: "생산",
        3: "개입",
        4: "품 명",
        5: "총 오더(팩)",
        6: "총 오더(pcs)",
        7: "용마 총입고수량(팩)",
        8: "용마 총입고수량(pcs)",
        9: "총 오더 잔량(팩)",
        10: "오더 잔량 (pcs)",
        11: "오더 낱개 총량",
        12: "",
        13: "",
        14: "낱개 재고(본사)",
        15: "생산 필요량 (pcs)",
        16: "배치 비율",
        17: "매칭 낱개 포장계획",
        18: "포장완료비율",
        19: "포장잔량비율",
        20: "주간포장계획",
        21: "주간 포장진도율",
        22: "총 포장진도율",
    }
    for col_idx, label in master_headers.items():
        ws.cell(row=header_row, column=col_idx, value=label)

    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="111827")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, size=11)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx in range(1, 36):  # A~AI
        cell = ws.cell(row=header_row, column=col_idx)
        if col_idx <= 22:
            cell.fill = header_fill
            cell.font = header_font
        cell.alignment = center
        cell.border = border

    widths = {
        1: 11,
        2: 11,
        3: 8,
        4: 34,
        5: 13,
        6: 14,
        7: 19,
        8: 19,
        9: 13,
        10: 14,
        11: 14,
        12: 12,
        13: 13,
        14: 14,
        15: 16,
        16: 10,
        17: 17,
        18: 13,
        19: 13,
        20: 13,
        21: 14,
        22: 14,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    for i, row in src.iterrows():
        r = data_start_row + i
        ws.cell(r, 1, row["제품코드"])
        ws.cell(r, 2, row["P코드"])
        ws.cell(r, 3, row["PACK당낱개수"])
        ws.cell(r, 4, row["품명"])
        ws.cell(r, 5, to_int_or_none(row["요청수량_PACK"]))
        ws.cell(r, 6, to_int_or_none(row["요청수량_낱개"]))
        ws.cell(r, 7, to_int_or_none(row["매칭출고수량_EA"]))
        ws.cell(r, 8, to_int_or_none(row["출고수량_낱개"]))
        ws.cell(r, 9, to_int_or_none(row["잔량"]))
        ws.cell(r, 10, to_int_or_none(row["잔량_낱개"]))
        ws.cell(r, 11, group_sum_j.get(row["_group_key"], 0))
        ws.cell(r, 12, None)
        ws.cell(r, 13, None)
        ws.cell(r, 14, to_int_or_none(row["_pcode_stock_total"]))
        ws.cell(r, 15, group_need_m.get(row["_group_key"], 0))

        for c in range(1, 36):  # A~AI
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = left if c == 4 else center
            if c in [5, 6, 7, 8, 9, 10, 11, 14, 15]:
                cell.number_format = "#,##0"
        ws.cell(r, 4).font = bold_font

    # 같은 P코드끼리 K(오더낱개총량), N(낱개재고), O(생산필요)만 병합 후
    # L/M 삭제를 거치면 최종 K/L/M만 병합 상태가 된다.
    pcode_group_ranges: list[tuple[int, int, str]] = []
    if not src.empty:
        merge_cols = [11, 14, 15]
        group_start = data_start_row
        prev_key = src.loc[0, "_group_key"]
        for i in range(1, len(src) + 1):
            is_end = i == len(src)
            curr_key = None if is_end else src.loc[i, "_group_key"]
            if is_end or curr_key != prev_key:
                start_row = group_start
                end_row = data_start_row + i - 1
                if end_row > start_row:
                    for col in merge_cols:
                        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=col, end_column=col)
                top_k = ws.cell(start_row, 11)
                top_k.value = group_sum_j.get(prev_key, 0)
                top_k.number_format = "#,##0"
                top_k.alignment = center
                for col in [12, 13, 15]:
                    cell = ws.cell(start_row, col)
                    cell.value = None
                    cell.alignment = center
                top_n = ws.cell(start_row, 14)
                if str(prev_key).startswith("P:"):
                    top_n.value = to_int_or_none(group_stock_l.get(str(prev_key), 0.0))
                    top_n.number_format = "#,##0"
                else:
                    top_n.value = None
                top_n.alignment = center
                top_o = ws.cell(start_row, 15)
                top_o.value = group_need_m.get(prev_key, 0)
                top_o.number_format = "#,##0"
                top_o.alignment = center
                if str(prev_key).startswith("P:"):
                    pcode_group_ranges.append((start_row, end_row, str(prev_key)))
                if not is_end:
                    group_start = data_start_row + i
                    prev_key = curr_key

    last_data_row = data_start_row + len(src) - 1 if not src.empty else data_start_row

    # A~V 전체 셀 기본 테두리 적용(헤더+데이터)
    for row in range(header_row, last_data_row + 1):
        for col in range(1, 23):  # A~V
            cell = ws.cell(row, col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # P코드 그룹별 바깥쪽 굵은 테두리(A~V)
    for start_row, end_row, _ in pcode_group_ranges:
        for row in range(start_row, end_row + 1):
            for col in range(1, 23):  # A~V
                cell = ws.cell(row, col)
                b = cell.border
                left_side = thick if col == 1 else b.left
                right_side = thick if col == 22 else b.right
                top_side = thick if row == start_row else b.top
                bottom_side = thick if row == end_row else b.bottom
                cell.border = Border(
                    left=left_side,
                    right=right_side,
                    top=top_side,
                    bottom=bottom_side,
                )

    # L(완제품 재고), M(포장공정재고) 열은 출력에서 완전히 제거
    ws.delete_cols(12, 2)
    # 최종 시트 기준 K/L/M 병합을 재적용
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col <= 13 and merged_range.max_col >= 11:
            ws.unmerge_cells(str(merged_range))
    for start_row, end_row, group_key in pcode_group_ranges:
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=11, end_column=11)  # K
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=12, end_column=12)  # L
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=13, end_column=13)  # M
        top_k_final = ws.cell(start_row, 11)
        top_k_final.value = group_sum_j.get(group_key, 0)
        top_k_final.number_format = "#,##0"
        top_k_final.alignment = center
        top_l_final = ws.cell(start_row, 12)
        top_l_final.value = to_int_or_none(group_stock_l.get(group_key, 0.0))
        top_l_final.number_format = "#,##0"
        top_l_final.alignment = center
        top_m_final = ws.cell(start_row, 13)
        top_m_final.value = group_need_m.get(group_key, 0)
        top_m_final.number_format = "#,##0"
        top_m_final.alignment = center
    # 최종 시트 기준 N열(14) 이후는 병합하지 않음
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col >= 14:
            ws.unmerge_cells(str(merged_range))
    # N/O 열 테두리를 P열과 동일하게 맞춤
    for row in range(header_row, last_data_row + 1):
        border_ref = ws.cell(row, 16).border  # P열
        ws.cell(row, 14).border = Border(
            left=border_ref.left,
            right=border_ref.right,
            top=border_ref.top,
            bottom=border_ref.bottom,
        )
        ws.cell(row, 15).border = Border(
            left=border_ref.left,
            right=border_ref.right,
            top=border_ref.top,
            bottom=border_ref.bottom,
        )

    ws.freeze_panes = "A4"
    ws.row_dimensions[header_row].height = 22


    # 세부사항 포장계획 입력 시트 (요청수량 열 포함)
    # 1번 시트(요약)에서 생산 필요량(M열) 제거
    ws.delete_cols(13, 1)

    input_sheet_title = "\uc138\ubd80\uc0ac\ud56d_\ud3ec\uc7a5\uacc4\ud68d\uc785\ub825"
    input_ws = wb.create_sheet(title=input_sheet_title)
    plan_days = 14
    capped_done_pack_col = 14
    capped_done_col = 15
    capped_progress_col = 16
    progress_col = 17
    weekly_plan_col = 18
    first_plan_col = 19
    last_plan_col = first_plan_col + plan_days - 1
    owner_col = last_plan_col + 1
    memo_col = owner_col + 1
    last_input_col = memo_col
    input_ws.merge_cells(f"A1:{get_column_letter(last_input_col)}1")
    input_ws["A1"] = "\ub178\ub780\uc0c9 \uc140\ub9cc \uc785\ub825: 14\uc77c \ud3ec\uc7a5\uacc4\ud68d(pcs), \ub2f4\ub2f9\uc790, \uba54\ubaa8"
    input_ws["A1"].font = Font(bold=True, color="1E3A8A")
    input_ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    input_headers = {
        1: "\ub9c8\uc2a4\ud130P\ucf54\ub4dc",
        2: "\ub300\ud45c\uc81c\ud488\ucf54\ub4dc > \ud310\ub9e4\ucf54\ub4dc",
        3: "\uac1c\uc785\uc218\ub7c9",
        4: "\ub300\ud45c\uc81c\ud488\uba85",
        5: "S코드",
        6: "P코드",
        7: "\uc694\uccad\uc218\ub7c9(pack)",
        8: "\uc694\uccad\uc218\ub7c9(pcs)",
        9: "\uc644\ub8cc\uc218\ub7c9(pack)",
        10: "\uc644\ub8cc\uc218\ub7c9(pcs)",
        11: "\ud604\uc7ac\uace0(pcs)",
        12: "\ud3ec\uc7a5\ud544\uc694\uc218\ub7c9(pcs)",
        13: "\ud3ec\uc7a5\uac00\ub2a5\uc218\ub7c9(pcs)",
        14: "초과수량 제외 완료 수량(pack)",
        15: "초과수량 제외 완료 수량(pcs)",
        16: "초과수량제외 포장진도율(%)",
        17: "\ucd1d \ud3ec\uc7a5\uc9c4\ub3c4\uc728(%)",
        18: "\uc8fc\uac04 \ud3ec\uc7a5 \uc218\ub7c9",
    }
    today = datetime.now().date()
    for day_offset in range(plan_days):
        col_idx = first_plan_col + day_offset
        input_headers[col_idx] = (today + timedelta(days=day_offset)).strftime("%Y-%m-%d")
    input_headers[owner_col] = "\ub2f4\ub2f9\uc790"
    input_headers[memo_col] = "\uba54\ubaa8"
    for col_idx, label in input_headers.items():
        cell = input_ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    input_ws.column_dimensions["A"].width = 20
    input_ws.column_dimensions["B"].width = 24
    input_ws.column_dimensions["C"].width = 10
    input_ws.column_dimensions["D"].width = 30
    input_ws.column_dimensions["E"].width = 16
    input_ws.column_dimensions["F"].width = 20
    input_ws.column_dimensions["G"].width = 16
    input_ws.column_dimensions["H"].width = 16
    input_ws.column_dimensions["I"].width = 16
    input_ws.column_dimensions["J"].width = 16
    input_ws.column_dimensions["K"].width = 16
    input_ws.column_dimensions["L"].width = 16
    input_ws.column_dimensions["M"].width = 16
    input_ws.column_dimensions["N"].width = 22
    input_ws.column_dimensions["O"].width = 22
    input_ws.column_dimensions["P"].width = 22
    input_ws.column_dimensions["Q"].width = 18
    input_ws.column_dimensions["R"].width = 16
    for c in range(first_plan_col, last_plan_col + 1):
        input_ws.column_dimensions[get_column_letter(c)].width = 13
    input_ws.column_dimensions[get_column_letter(owner_col)].width = 14
    input_ws.column_dimensions[get_column_letter(memo_col)].width = 32

    def n_int(v: object) -> int:
        n = pd.to_numeric(pd.Series([v]), errors="coerce").iloc[0]
        if pd.isna(n):
            return 0
        return int(round(float(n)))

    def n_float(v: object) -> float:
        n = pd.to_numeric(pd.Series([v]), errors="coerce").iloc[0]
        if not pd.isna(n):
            return float(n)
        m = re.search(r"(\d+(?:\.\d+)?)", str(v))
        return float(m.group(1)) if m else 0.0

    def derive_s_code(rep_code_value: object, p_code_value: object, raw_s_code: object = "") -> str:
        raw_text = normalize_key_value(raw_s_code)
        if raw_text and "-" in raw_text:
            return raw_text
        rep_text = normalize_key_value(rep_code_value)
        p_text = normalize_key_value(p_code_value)
        if rep_text and p_text:
            m = re.search(r"-(\d+(?:\.\d+)?)", p_text)
            if m:
                return f"{rep_text}-{m.group(1)}"
        return raw_text or rep_text

    pcode_detail_map = pcode_detail_map or {}
    detail_codes_from_item: dict[str, list[str]] = {}
    if detail_item_df is not None and not detail_item_df.empty:
        item_pcode_col = find_col(detail_item_df, ["P코드", "P 코드"])
        item_detail_col = find_col(detail_item_df, ["제품코드(마스터)", "세부코드"])
        if item_pcode_col and item_detail_col:
            for _, irow in detail_item_df.iterrows():
                detail_code = normalize_key_value(irow.get(item_detail_col, ""))
                if not detail_code:
                    continue
                for m in split_master_codes(irow.get(item_pcode_col, "")):
                    if m not in detail_codes_from_item:
                        detail_codes_from_item[m] = []
                    if detail_code not in detail_codes_from_item[m]:
                        detail_codes_from_item[m].append(detail_code)
    pcode_stock_totals = pcode_stock_totals or {}
    detail_stock_by_code = detail_stock_by_code or {}
    detail_rows: dict[tuple[str, str, str], dict[str, object]] = {}
    pcode_restore_log: dict[str, dict[str, object]] = {}
    detail_item_has_metrics = False
    if detail_item_df is not None and not detail_item_df.empty:
        _pcol = find_col(detail_item_df, ["P코드", "P 코드"])
        _dcol = find_col(detail_item_df, ["제품코드(마스터)", "세부코드"])
        _need = find_col(detail_item_df, ["잔량_낱개"])
        _done = find_col(detail_item_df, ["출고수량_낱개"])
        detail_item_has_metrics = bool(_pcol and _dcol and (_need or _done))

    for src_order, (_, row) in enumerate(src.iterrows()):
        rep_code = normalize_key_value(row.get(code_col, ""))
        rep_name = normalize_key_value(row.get(name_col, ""))
        req_total = n_int(row.get(required_cols[5], 0))  # 요청수량_낱개
        need_total = n_int(row.get(required_cols[9], 0))  # 잔량_낱개
        done_total = n_int(row.get(required_cols[7], 0))  # 출고수량_낱개
        raw_pcode_text = row.get(pcode_col, "")
        base_masters = split_master_codes(raw_pcode_text)
        masters = expanded_master_codes(raw_pcode_text, row.get(code_col, ""))
        if rep_code:
            added = max(len(masters) - len(base_masters), 0)
            if added > 0:
                prev = pcode_restore_log.get(rep_code)
                if prev is None or int(prev.get("추가코드수", 0)) < added:
                    pcode_restore_log[rep_code] = {
                        "제품코드": rep_code,
                        "품명": rep_name,
                        "요약P코드수": len(base_masters),
                        "복원P코드수": len(masters),
                        "추가코드수": added,
                        "요약P코드": ", ".join(base_masters),
                    }
        if not masters:
            continue
        m_n = max(1, len(masters))
        m_req_base, m_req_rem = divmod(req_total, m_n)
        m_need_base, m_need_rem = divmod(need_total, m_n)
        m_done_base, m_done_rem = divmod(done_total, m_n)
        for m_idx, master_code in enumerate(masters):
            m_req = m_req_base + (1 if m_idx < m_req_rem else 0)
            m_need = m_need_base + (1 if m_idx < m_need_rem else 0)
            m_done = m_done_base + (1 if m_idx < m_done_rem else 0)
            details = detail_codes_from_item.get(master_code, [])
            if not details:
                details = split_master_codes(pcode_detail_map.get(master_code, ""))
            if not details:
                details = [master_code]
            d_n = max(1, len(details))
            d_req_base, d_req_rem = divmod(m_req, d_n)
            d_need_base, d_need_rem = divmod(m_need, d_n)
            d_done_base, d_done_rem = divmod(m_done, d_n)
            for d_idx, detail_code in enumerate(details):
                key = (master_code, detail_code, rep_code)
                item = detail_rows.setdefault(
                    key,
                    {
                        "pcode": detail_code,
                        "master": master_code,
                        "rep_code": rep_code,
                        "s_code": "",
                        "pack_size": normalize_key_value(row.get(required_cols[2], "")),
                        "rep_name": rep_name,
                        "order": src_order,
                        "stock": 0,
                        "requested": 0,
                        "done": 0,
                        "need": 0,
                    },
                )
                if src_order < int(item.get("order", src_order)):
                    item["order"] = src_order
                if not item["rep_code"] and rep_code:
                    item["rep_code"] = rep_code
                if not item.get("pack_size"):
                    item["pack_size"] = normalize_key_value(row.get(required_cols[2], ""))
                if not item["rep_name"] and rep_name:
                    item["rep_name"] = rep_name
                item["requested"] += d_req_base + (1 if d_idx < d_req_rem else 0)
                # detail_item_df 유무와 무관하게 기본 수량은 src 분배값으로 채운다.
                item["done"] += d_done_base + (1 if d_idx < d_done_rem else 0)
                item["need"] += d_need_base + (1 if d_idx < d_need_rem else 0)

    # item 상세 기준 보강: src/pcode_detail_map에서 누락된 세부코드도 강제 포함
    if detail_item_df is not None and not detail_item_df.empty:
        item_pcode_col2 = find_col(detail_item_df, ["P코드", "P 코드"])
        item_detail_col2 = find_col(detail_item_df, ["제품코드(마스터)", "세부코드"])
        item_rep_col2 = find_col(detail_item_df, ["제품코드", "품목코드"])
        item_scode_col2 = find_col(detail_item_df, ["품목코드", "제품코드"])
        item_pack_col2 = find_col(detail_item_df, ["PACK당낱개수"])
        item_name_col2 = find_col(detail_item_df, ["품명", "품명_요청", "품명_출고"])
        item_req_col2 = find_col(detail_item_df, [required_cols[5], "요청수량_낱개"])
        item_need_col2 = find_col(detail_item_df, ["잔량_낱개"])
        item_done_col2 = find_col(detail_item_df, ["출고수량_낱개"])
        if item_pcode_col2 and item_detail_col2:
            for _, irow in detail_item_df.iterrows():
                detail_code = normalize_key_value(irow.get(item_detail_col2, ""))
                if not detail_code:
                    continue
                rep_code2 = normalize_key_value(irow.get(item_rep_col2, "")) if item_rep_col2 else ""
                s_code2 = normalize_key_value(irow.get(item_scode_col2, "")) if item_scode_col2 else rep_code2
                pack_size2 = normalize_key_value(irow.get(item_pack_col2, "")) if item_pack_col2 else ""
                rep_name2 = normalize_key_value(irow.get(item_name_col2, "")) if item_name_col2 else ""
                masters = split_master_codes(irow.get(item_pcode_col2, ""))
                if not masters:
                    continue
                m_n = max(1, len(masters))
                for m_idx, m in enumerate(masters):
                    key = (m, detail_code, rep_code2)
                    created_now = key not in detail_rows
                    row_obj = detail_rows.setdefault(
                        key,
                        {
                            "pcode": detail_code,
                            "master": m,
                            "rep_code": rep_code2,
                            "s_code": s_code2,
                            "pack_size": pack_size2,
                            "rep_name": rep_name2,
                            "order": 10**9,
                            "stock": 0,
                            "requested": 0,
                            "done": 0,
                            "need": 0,
                        },
                    )
                    if not row_obj["rep_code"] and rep_code2:
                        row_obj["rep_code"] = rep_code2
                    if s_code2 and ("-" in s_code2 or not row_obj.get("s_code")):
                        row_obj["s_code"] = s_code2
                    if not row_obj.get("pack_size") and pack_size2:
                        row_obj["pack_size"] = pack_size2
                    if not row_obj["rep_name"] and rep_name2:
                        row_obj["rep_name"] = rep_name2
                    # 수량(requested/done/need)은 원본 src 분배값만 사용하고, detail_item_df 보정 단계에서는 덮어쓰지 않음

    # S코드(품목코드) 기준 요청수량(pack) 맵 생성: 1분기 요청수량 우선 참조
    request_pack_by_s_code: dict[str, float] = {}
    allowed_s_codes: set[str] = set()
    if raw_req_df is not None and not raw_req_df.empty:
        req_s_col = find_col(raw_req_df, ["품목코드", "S코드", "제품코드"])
        req_q1_pack_col = find_col(raw_req_df, ["1분기요청수량", "1분기 요청수량", "요청수량_PACK", "분기요청수량"])
        if req_s_col:
            for _, rr in raw_req_df.iterrows():
                s_key = normalize_key_value(rr.get(req_s_col, ""))
                if not s_key:
                    continue
                allowed_s_codes.add(s_key)
                req_pack = n_float(rr.get(req_q1_pack_col, 0)) if req_q1_pack_col else 0.0
                request_pack_by_s_code[s_key] = float(request_pack_by_s_code.get(s_key, 0.0) + req_pack)

    # S코드(품목코드) 기준 완료수량(pcs) 맵 생성: 입고수량 원본 동일 S코드는 합산
    done_pcs_by_s_code: dict[str, int] = {}
    if raw_inbound_df is not None and not raw_inbound_df.empty:
        in_s_col = find_col(raw_inbound_df, ["품목코드", "S코드", "제품코드"])
        in_qty_col = find_col(raw_inbound_df, ["출고수량_낱개", "출고수량_낱개환산", "출고수량_EA", "수량"])
        if in_s_col and in_qty_col:
            for _, ir in raw_inbound_df.iterrows():
                s_key = normalize_key_value(ir.get(in_s_col, ""))
                if not s_key:
                    continue
                done_pcs = n_int(ir.get(in_qty_col, 0))
                done_pcs_by_s_code[s_key] = int(done_pcs_by_s_code.get(s_key, 0) + done_pcs)

    # 현재고는 마스터 합계 분배가 아니라 세부 P코드 기준 재고 사용
    if detail_stock_by_code:
        for row_obj in detail_rows.values():
            row_obj["stock"] = n_int(detail_stock_by_code.get(str(row_obj.get("pcode", "")), 0.0))

    editable_fill = PatternFill("solid", fgColor="FEF3C7")
    sorted_rows = sorted(
        detail_rows.values(),
        key=lambda x: (
            int(x.get("order", 10**9)),  # MASTER 포장계획 양식과 동일한 원본 정렬 순서
            str(x["master"]),
            str(x["rep_code"]),
            str(x["pcode"]),
        ),
    )
    visible_rows = []
    for r in sorted_rows:
        if int(r.get("requested", 0)) + int(r.get("done", 0)) + int(r.get("need", 0)) <= 0:
            continue
        s_code_key = derive_s_code(r.get("rep_code", ""), r.get("pcode", ""), r.get("s_code", ""))
        # 원본 생산요청리스트에 존재하는 S코드만 내보낸다.
        if allowed_s_codes and s_code_key not in allowed_s_codes:
            continue
        visible_rows.append(r)
    alloc_by_obj_id: dict[int, int] = {}
    pcode_groups: dict[str, list[dict[str, object]]] = {}
    for row_obj in visible_rows:
        pkey = str(row_obj.get("pcode", ""))
        pcode_groups.setdefault(pkey, []).append(row_obj)
    for pkey, rows in pcode_groups.items():
        remaining = n_int(detail_stock_by_code.get(pkey, 0.0))
        rows_sorted = sorted(
            rows,
            key=lambda r: (
                -pack_sort_value(r.get("pack_size", "")),  # 개입수량 큰 순서 우선 배분
                int(r.get("order", 10**9)),
            ),
        )
        for row_obj in rows_sorted:
            need_qty = int(row_obj.get("need", 0))
            alloc = max(0, min(need_qty, remaining))
            alloc_by_obj_id[id(row_obj)] = alloc
            remaining -= alloc

    for i, info in enumerate(visible_rows):
        r = data_start_row + i
        requested_qty_raw = int(info.get("requested", 0))
        s_code_value = derive_s_code(info.get("rep_code", ""), info.get("pcode", ""), info.get("s_code", ""))
        # 입고 원본(S코드)에서 확인되지 않은 완료수량은 0으로 처리한다.
        done_qty = int(done_pcs_by_s_code.get(s_code_value, 0))
        need_qty = int(info["need"])
        pack_size_num = n_float(info.get("pack_size", 0))
        if pack_size_num <= 0:
            pack_size_num = 1.0
        req_pack_by_s_code = float(request_pack_by_s_code.get(s_code_value, 0.0))
        # 요청수량(pack)은 S코드 기준 1분기 요청수량을 우선 사용
        requested_pack = req_pack_by_s_code if req_pack_by_s_code > 0 else (float(requested_qty_raw) / pack_size_num if requested_qty_raw > 0 else 0.0)
        requested_qty = int(round(requested_pack * pack_size_num)) if requested_pack > 0 else (requested_qty_raw if requested_qty_raw > 0 else (done_qty + need_qty))
        done_pack = float(done_qty) / pack_size_num if pack_size_num > 0 else 0.0
        # 포장필요수량은 요청수량(pcs) - 완료수량(pcs)
        need_qty = int(max(requested_qty - done_qty, 0))
        packable_qty = int(alloc_by_obj_id.get(id(info), 0))
        input_ws.cell(r, 1, info["master"])
        input_ws.cell(r, 2, info["rep_code"])
        input_ws.cell(r, 3, info.get("pack_size", ""))
        input_ws.cell(r, 4, info["rep_name"])
        input_ws.cell(r, 5, s_code_value)
        input_ws.cell(r, 6, info["pcode"])
        input_ws.cell(r, 7, requested_pack)
        input_ws.cell(r, 8, f"=IFERROR(G{r}*C{r},0)")
        input_ws.cell(r, 9, done_pack)
        input_ws.cell(r, 10, f"=IFERROR(I{r}*C{r},0)")
        input_ws.cell(r, 11, int(info["stock"]))
        input_ws.cell(r, 12, need_qty)
        input_ws.cell(r, 13, packable_qty)
        capped_done_qty = min(done_qty, requested_qty)
        capped_done_pack = float(capped_done_qty) / pack_size_num if pack_size_num > 0 else 0.0
        input_ws.cell(r, capped_done_pack_col, capped_done_pack)
        input_ws.cell(r, capped_done_col, capped_done_qty)
        input_ws.cell(r, capped_progress_col, f"=IFERROR({get_column_letter(capped_done_col)}{r}/H{r},0)")
        input_ws.cell(r, progress_col, f"=IFERROR(J{r}/H{r},0)")
        input_ws.cell(
            r,
            weekly_plan_col,
            f"=SUM({get_column_letter(first_plan_col)}{r}:{get_column_letter(last_plan_col)}{r})",
        )
        for c in range(first_plan_col, last_plan_col + 1):
            input_ws.cell(r, c, None)
        input_ws.cell(r, owner_col, None)
        input_ws.cell(r, memo_col, None)
        for c in range(1, last_input_col + 1):
            cell = input_ws.cell(r, c)
            cell.border = border
            cell.alignment = left if c in [4, memo_col] else center
            if c in [3, 7, 8, 9, 10, 11, 12, 13, capped_done_pack_col, capped_done_col, weekly_plan_col] or (first_plan_col <= c <= last_plan_col):
                cell.number_format = "#,##0"
            elif c in [capped_progress_col, progress_col]:
                cell.number_format = "0.0%"
        for c in list(range(first_plan_col, last_plan_col + 1)) + [owner_col, memo_col]:
            input_ws.cell(r, c).fill = editable_fill

    # A~D열에서 같은 값이 연속되는 구간은 병합
    if visible_rows:
        last_row = data_start_row + len(visible_rows) - 1
        input_ws.conditional_formatting.add(
            f"{get_column_letter(capped_progress_col)}{data_start_row}:{get_column_letter(capped_progress_col)}{last_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color="4F81BD",
                showValue=True,
            ),
        )
        input_ws.conditional_formatting.add(
            f"{get_column_letter(progress_col)}{data_start_row}:{get_column_letter(progress_col)}{last_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color="4F81BD",
                showValue=True,
            ),
        )
        # 판매코드(B열)는 SUMIFS 집계를 위해 병합하지 않음
        for col in [1, 4]:
            start_row = data_start_row
            prev_val = str(input_ws.cell(start_row, col).value or "")
            for r in range(data_start_row + 1, last_row + 2):
                curr_val = str(input_ws.cell(r, col).value or "") if r <= last_row else "__END__"
                if curr_val != prev_val:
                    end_row = r - 1
                    if end_row > start_row and prev_val != "":
                        input_ws.merge_cells(
                            start_row=start_row,
                            end_row=end_row,
                            start_column=col,
                            end_column=col,
                        )
                        input_ws.cell(start_row, col).alignment = center
                    start_row = r
                    prev_val = curr_val

    # ???? ?? ?? ?? (????_?????? ??)
    sales_summary_ws = wb.create_sheet(title="판매코드 요약")
    sales_headers = {
        1: "판매코드",
        2: "개입수",
        3: "대표제품명",
        4: "요청수량(pack)",
        5: "요청수량(pcs)",
        6: "완료수량(pack)",
        7: "완료수량(pcs)",
        8: "포장필요수량(pack)",
        9: "포장필요수량(pcs)",
        10: "주간 포장 수량(pack)",
        11: "주간 포장 수량(pcs)",
        12: "현재 포장진도율(%)",
        13: "예상 포장진도율(%)",
    }
    for col_idx, label in sales_headers.items():
        cell = sales_summary_ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    sales_summary_ws.column_dimensions["A"].width = 20
    sales_summary_ws.column_dimensions["B"].width = 10
    sales_summary_ws.column_dimensions["C"].width = 30
    for c in range(4, 12):
        sales_summary_ws.column_dimensions[get_column_letter(c)].width = 16
    sales_summary_ws.column_dimensions["L"].width = 18
    sales_summary_ws.column_dimensions["M"].width = 26

    sales_code_to_name: dict[str, str] = {}
    sales_code_to_pack: dict[str, float] = {}

    def to_pack_size_number(value: object) -> float:
        n = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        if not pd.isna(n):
            return float(n)
        m = re.search(r"(\d+(?:\.\d+)?)", str(value))
        return float(m.group(1)) if m else 0.0

    for info in visible_rows:
        sales_code = normalize_key_value(info.get("rep_code", "")) or "(미지정)"
        rep_name = normalize_key_value(info.get("rep_name", ""))
        pack_size = to_pack_size_number(info.get("pack_size", 0))
        if sales_code not in sales_code_to_name:
            sales_code_to_name[sales_code] = rep_name
        elif not sales_code_to_name[sales_code] and rep_name:
            sales_code_to_name[sales_code] = rep_name
        sales_code_to_pack[sales_code] = max(float(sales_code_to_pack.get(sales_code, 0.0)), float(pack_size))

    input_ref = f"'{input_sheet_title}'"
    sales_codes = list(sales_code_to_name.keys())
    for idx, sales_code in enumerate(sales_codes):
        r = data_start_row + idx
        sales_summary_ws.cell(r, 1, sales_code)
        sales_summary_ws.cell(r, 2, sales_code_to_pack.get(sales_code, 0.0))
        sales_summary_ws.cell(r, 3, sales_code_to_name.get(sales_code, ""))

        sales_summary_ws.cell(r, 4, f"=SUMIFS({input_ref}!$G:$G,{input_ref}!$B:$B,$A{r})")
        sales_summary_ws.cell(r, 5, f"=SUMIFS({input_ref}!$H:$H,{input_ref}!$B:$B,$A{r})")
        sales_summary_ws.cell(
            r,
            6,
            f"=SUMIFS({input_ref}!${get_column_letter(capped_done_pack_col)}:${get_column_letter(capped_done_pack_col)},{input_ref}!$B:$B,$A{r})",
        )
        sales_summary_ws.cell(
            r,
            7,
            f"=SUMIFS({input_ref}!${get_column_letter(capped_done_col)}:${get_column_letter(capped_done_col)},{input_ref}!$B:$B,$A{r})",
        )
        sales_summary_ws.cell(r, 9, f"=SUMIFS({input_ref}!$L:$L,{input_ref}!$B:$B,$A{r})")
        sales_summary_ws.cell(r, 8, f"=IFERROR(I{r}/$B{r},0)")
        sales_summary_ws.cell(
            r,
            11,
            f"=SUMIFS({input_ref}!${get_column_letter(weekly_plan_col)}:${get_column_letter(weekly_plan_col)},{input_ref}!$B:$B,$A{r})",
        )
        sales_summary_ws.cell(r, 10, f"=IFERROR(K{r}/$B{r},0)")
        sales_summary_ws.cell(r, 12, f"=IFERROR(G{r}/E{r},0)")
        sales_summary_ws.cell(r, 13, f"=IFERROR((G{r}+K{r})/E{r},0)")

        for c in range(1, 14):
            cell = sales_summary_ws.cell(r, c)
            cell.border = border
            cell.alignment = left if c == 3 else center
            if c in [2, 4, 5, 6, 7, 8, 9, 10, 11]:
                cell.number_format = "#,##0"
            elif c in [12, 13]:
                cell.number_format = "0.0%"

    if sales_codes:
        last_sales_row = data_start_row + len(sales_codes) - 1
        sales_summary_ws.conditional_formatting.add(
            f"L{data_start_row}:L{last_sales_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color="4F81BD",
                showValue=True,
            ),
        )
        sales_summary_ws.conditional_formatting.add(
            f"M{data_start_row}:M{last_sales_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color="93C47D",
                showValue=True,
            ),
        )

    sales_summary_ws.freeze_panes = "A4"
    sales_summary_ws.row_dimensions[header_row].height = 22

    input_ws.freeze_panes = "A4"
    input_ws.row_dimensions[header_row].height = 22

    # 전체요약 시트 (코드 구분 없이 전체 합계/진도율)
    overall_ws = wb.create_sheet(title="전체요약")

    # B~G ??, A?? pack/pcs ??
    overall_headers = {
        2: "전체 요청수량",
        3: "전체 완료수량",
        4: "전체 포장필요수량",
        5: "전체 주간 포장 수량",
        6: "현재 포장진도율",
        7: "예상 포장진도율",
    }
    for col_idx, label in overall_headers.items():
        cell = overall_ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    overall_ws.cell(2, 1, "pack")
    overall_ws.cell(3, 1, "pcs")
    overall_ws.cell(2, 1).font = bold_font
    overall_ws.cell(3, 1).font = bold_font

    input_ref = f"'{input_sheet_title}'"

    # pack 행: 판매코드 요약 시트의 pack 컬럼 합계를 재사용
    summary_ref = f"'{sales_summary_ws.title}'"
    overall_ws.cell(2, 2, f"=SUM({summary_ref}!$D:$D)")   # 요청수량(pack)
    overall_ws.cell(2, 3, f"=SUM({summary_ref}!$F:$F)")   # 완료수량(pack)
    overall_ws.cell(2, 4, f"=SUM({summary_ref}!$H:$H)")   # 포장필요수량(pack)
    overall_ws.cell(2, 5, f"=SUM({summary_ref}!$J:$J)")   # 주간 포장 수량(pack)
    # 진도율은 pcs 기준과 동일하게 표시( pack/pcs 진도율 일치 )
    overall_ws.cell(2, 6, "=F3")
    overall_ws.cell(2, 7, "=G3")

    # pcs 행도 판매코드 요약 시트(초과수량 제외 완료수량 기준) 집계를 재사용
    overall_ws.cell(3, 2, f"=SUM({summary_ref}!$E:$E)")   # 요청수량(pcs)
    overall_ws.cell(3, 3, f"=SUM({summary_ref}!$G:$G)")   # 완료수량(pcs, 초과제외)
    overall_ws.cell(3, 4, f"=SUM({summary_ref}!$I:$I)")   # 포장필요수량(pcs)
    overall_ws.cell(3, 5, f"=SUM({summary_ref}!$K:$K)")   # 주간 포장 수량(pcs)
    overall_ws.cell(3, 6, "=IFERROR(C3/B3,0)")
    overall_ws.cell(3, 7, "=IFERROR((C3+E3)/B3,0)")

    overall_ws.column_dimensions["A"].width = 16
    overall_ws.column_dimensions["B"].width = 14
    overall_ws.column_dimensions["C"].width = 14
    overall_ws.column_dimensions["D"].width = 16
    overall_ws.column_dimensions["E"].width = 16
    overall_ws.column_dimensions["F"].width = 14
    overall_ws.column_dimensions["G"].width = 14

    for r in [2, 3]:
        for c in range(1, 8):
            cell = overall_ws.cell(r, c)
            cell.border = border
            cell.alignment = center
            if c in [2, 3, 4, 5]:
                cell.number_format = "#,##0"
            elif c in [6, 7]:
                cell.number_format = "0.0%"

    overall_ws.conditional_formatting.add(
        "F2:F3",
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color="4F81BD",
            showValue=True,
        ),
    )
    overall_ws.conditional_formatting.add(
        "G2:G3",
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color="93C47D",
            showValue=True,
        ),
    )
    overall_ws.freeze_panes = "A2"
    overall_ws.row_dimensions[1].height = 22

    # P코드 복원 로그 시트
    # 초기 생성용 MASTER 시트는 최종 파일에서 제외
    wb.remove(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_color_metric(
    container,
    label: str,
    value: str,
    bg_color: str,
    border_color: str,
    label_color: str = "#1e3a8a",
    value_color: str = "#0f172a",
) -> None:
    container.markdown(
        f"""
        <div style="
            background:{bg_color};
            border:1px solid {border_color};
            border-radius:12px;
            padding:10px 14px;
            box-shadow:0 2px 8px rgba(37, 99, 235, 0.08);
            min-height:98px;
        ">
            <div style="font-size:0.86rem;color:{label_color};font-weight:700;">{label}</div>
            <div style="font-size:1.65rem;color:{value_color};font-weight:700;line-height:1.2;margin-top:6px;">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def apply_chart_style(chart: alt.Chart) -> alt.Chart:
    return (
        chart.configure_axis(
            labelFont=APP_FONT_FAMILY,
            titleFont=APP_FONT_FAMILY,
            labelFontSize=13,
            titleFontSize=15,
        )
        .configure_legend(
            labelFont=APP_FONT_FAMILY,
            titleFont=APP_FONT_FAMILY,
            labelFontSize=13,
            titleFontSize=14,
        )
        .configure_title(
            font=APP_FONT_FAMILY,
            fontSize=20,
        )
    )


try:
    req_raw, inbound_raw, req_file, inbound_file = load_data(".")
    item_product_master = load_item_product_master_map(".")
    req = prepare_request(req_raw)
    inbound = prepare_inbound(inbound_raw)
except Exception as e:
    st.error(f"데이터 로딩 실패: {e}")
    st.stop()

# 사이드바 필터
all_years = sorted(pd.concat([req["년"].dropna(), inbound["년"].dropna()]).astype(int).unique().tolist())
all_quarters = sorted(pd.concat([req["분기"].dropna(), inbound["분기"].dropna()]).astype(int).unique().tolist())

year_options = ["전체"] + all_years
quarter_options = ["전체"] + all_quarters
type_options = ["전체", "정규품", "샘플"]

st.sidebar.header("조회 조건")
selected_year = st.sidebar.selectbox("년", year_options, index=1 if len(year_options) > 1 else 0)
selected_quarter = st.sidebar.selectbox("분기", quarter_options, index=1 if len(quarter_options) > 1 else 0)
selected_type = st.sidebar.selectbox("구분", type_options, index=1)

st.sidebar.markdown("---")
st.sidebar.write(f"요청 파일: `{req_file}`")
st.sidebar.write(f"입고 파일: `{inbound_file}`")

req_f = req.copy()
in_f = inbound.copy()
if selected_year != "전체":
    req_f = req_f[req_f["년"] == selected_year]
    in_f = in_f[in_f["년"] == selected_year]
if selected_quarter != "전체":
    req_f = req_f[req_f["분기"] == selected_quarter]
    in_f = in_f[in_f["분기"] == selected_quarter]
if selected_type != "전체":
    req_f = req_f[req_f["구분"] == selected_type]
    req_keys = req_f[["년", "분기", "품목코드"]].drop_duplicates()
    if req_keys.empty:
        in_f = in_f.iloc[0:0].copy()
    else:
        in_f = in_f.merge(req_keys.assign(_keep=1), on=["년", "분기", "품목코드"], how="inner").drop(columns=["_keep"])

# P코드 고정 개입수는 요청/입고 모두 동일 품목코드에 강제 적용
for pcode, forced_pack in PACK_COUNT_PCODE_OVERRIDES.items():
    pcode_mask = req_f["P코드"].fillna("").astype(str).str.contains(
        rf"(^|,\s*){re.escape(pcode)}(\s*,|$)",
        regex=True,
    )
    forced_items = set(req_f.loc[pcode_mask, "품목코드"].fillna("").astype(str).str.strip())
    if forced_items:
        req_f.loc[req_f["품목코드"].fillna("").astype(str).str.strip().isin(forced_items), "PACK당낱개수"] = float(forced_pack)
        in_f.loc[in_f["품목코드"].fillna("").astype(str).str.strip().isin(forced_items), "PACK당낱개수"] = float(forced_pack)

# PACK 기준 수량을 낱개 기준으로 환산
req_f["요청수량_낱개환산"] = req_f["요청수량_PACK"] * req_f["PACK당낱개수"]
in_f["출고수량_낱개환산"] = in_f["출고수량_EA"] * in_f["PACK당낱개수"]

# 제품코드(앞4자리) 기준 요약
req_prod = req_f.groupby(["년", "분기", "제품코드"], as_index=False)["요청수량_PACK"].sum()
in_prod = in_f.groupby(["년", "분기", "제품코드"], as_index=False)["출고수량_EA"].sum()
prod = req_prod.merge(in_prod, on=["년", "분기", "제품코드"], how="outer").fillna(0)

pcode_prod = (
    req_f.groupby(["년", "분기", "제품코드"], as_index=False)["P코드"]
    .apply(summarize_codes)
    .rename(columns={"P코드": "P코드"})
)
prod = prod.merge(pcode_prod, on=["년", "분기", "제품코드"], how="left")
prod["P코드"] = prod["P코드"].fillna("")

brand_prod = (
    req_f.groupby(["년", "분기", "제품코드"], as_index=False)["브랜드"]
    .apply(summarize_codes)
    .rename(columns={"브랜드": "브랜드"})
)
prod = prod.merge(brand_prod, on=["년", "분기", "제품코드"], how="left")
prod["브랜드"] = prod["브랜드"].fillna("").astype(str).replace("0", "")
type_prod = (
    req_f.groupby(["년", "분기", "제품코드"], as_index=False)["구분"]
    .apply(summarize_codes)
    .rename(columns={"구분": "구분"})
)
prod = prod.merge(type_prod, on=["년", "분기", "제품코드"], how="left")
prod["구분"] = prod["구분"].fillna("").astype(str).replace("0", "")

# 제품명코드 요약의 품명은 요청 파일 기준으로만 표시
name_src = req_f[["년", "분기", "제품코드", "품명"]].copy()
prod_names = (
    name_src.groupby(["년", "분기", "제품코드"], as_index=False)["품명"]
    .apply(summarize_names)
    .rename(columns={"품명": "품명"})
)
prod = prod.merge(prod_names, on=["년", "분기", "제품코드"], how="left")
prod["품명"] = prod["품명"].fillna("")

pack_src_prod = pd.concat(
    [
        req_f[["년", "분기", "제품코드", "PACK당낱개수"]],
        in_f[["년", "분기", "제품코드", "PACK당낱개수"]],
    ],
    ignore_index=True,
)
prod_pack = (
    pack_src_prod.groupby(["년", "분기", "제품코드"], as_index=False)["PACK당낱개수"]
    .apply(summarize_pack_counts)
    .rename(columns={"PACK당낱개수": "PACK당낱개수"})
)
prod = prod.merge(prod_pack, on=["년", "분기", "제품코드"], how="left")
prod["PACK당낱개수"] = prod["PACK당낱개수"].fillna("")
for pcode, forced_pack in PACK_COUNT_PCODE_OVERRIDES.items():
    mask = prod["P코드"].fillna("").astype(str).str.contains(rf"(^|,\s*){re.escape(pcode)}(\s*,|$)", regex=True)
    prod.loc[mask, "PACK당낱개수"] = str(int(forced_pack))

req_piece_prod = (
    req_f.groupby(["년", "분기", "제품코드"], as_index=False)["요청수량_낱개환산"]
    .sum(min_count=1)
    .rename(columns={"요청수량_낱개환산": "요청수량_낱개"})
)
in_piece_prod = (
    in_f.groupby(["년", "분기", "제품코드"], as_index=False)["출고수량_낱개환산"]
    .sum(min_count=1)
    .rename(columns={"출고수량_낱개환산": "출고수량_낱개"})
)
prod = prod.merge(req_piece_prod, on=["년", "분기", "제품코드"], how="left")
prod = prod.merge(in_piece_prod, on=["년", "분기", "제품코드"], how="left")
prod_req_piece = pd.to_numeric(prod["요청수량_낱개"], errors="coerce").fillna(0)
prod_ship_piece = pd.to_numeric(prod["출고수량_낱개"], errors="coerce").fillna(0)
prod["매칭출고수량_낱개"] = np.minimum(prod_req_piece, prod_ship_piece)
prod["초과출고수량_낱개"] = np.maximum(prod_ship_piece - prod_req_piece, 0)
prod["잔량_낱개"] = np.maximum(prod_req_piece - prod_ship_piece, 0)

prod = add_progress_columns(prod, "요청수량_PACK", "출고수량_EA")

# 품목코드 상세 요약
req_item = req_f.groupby(["년", "분기", "품목코드"], as_index=False).agg(
    품명=("품명", "first"), 요청수량_PACK=("요청수량_PACK", "sum")
)
in_item = in_f.groupby(["년", "분기", "품목코드"], as_index=False).agg(
    품명=("품명", "first"), 출고수량_EA=("출고수량_EA", "sum")
)
item = req_item.merge(in_item, on=["년", "분기", "품목코드"], how="outer", suffixes=("_요청", "_출고")).fillna(0)
item["품명"] = np.where(item["품명_요청"].astype(str) != "0", item["품명_요청"], item["품명_출고"])
item["품명검색"] = item.apply(lambda r: merge_text_values(r["품명_요청"], r["품명_출고"]), axis=1)
item["제품군명"] = item["품명검색"].apply(normalize_product_family_name)
item["제품코드"] = item["품목코드"].astype(str).str[:4]
pcode_item = (
    req_f.groupby(["년", "분기", "품목코드"], as_index=False)["P코드"]
    .apply(summarize_codes)
    .rename(columns={"P코드": "P코드"})
)
item = item.merge(pcode_item, on=["년", "분기", "품목코드"], how="left")
item["P코드"] = item["P코드"].fillna("")

brand_item = (
    req_f.groupby(["년", "분기", "품목코드"], as_index=False)["브랜드"]
    .apply(summarize_codes)
    .rename(columns={"브랜드": "브랜드"})
)
item = item.merge(brand_item, on=["년", "분기", "품목코드"], how="left")
item["브랜드"] = item["브랜드"].fillna("").astype(str).replace("0", "")
type_item = (
    req_f.groupby(["년", "분기", "품목코드"], as_index=False)["구분"]
    .apply(summarize_codes)
    .rename(columns={"구분": "구분"})
)
item = item.merge(type_item, on=["년", "분기", "품목코드"], how="left")
item["구분"] = item["구분"].fillna("").astype(str).replace("0", "")

pack_src_item = pd.concat(
    [
        req_f[["년", "분기", "품목코드", "PACK당낱개수"]],
        in_f[["년", "분기", "품목코드", "PACK당낱개수"]],
    ],
    ignore_index=True,
)
item_pack = (
    pack_src_item.groupby(["년", "분기", "품목코드"], as_index=False)["PACK당낱개수"]
    .apply(summarize_pack_counts)
    .rename(columns={"PACK당낱개수": "PACK당낱개수"})
)
item = item.merge(item_pack, on=["년", "분기", "품목코드"], how="left")
item["PACK당낱개수"] = item["PACK당낱개수"].fillna("")
for pcode, forced_pack in PACK_COUNT_PCODE_OVERRIDES.items():
    mask = item["P코드"].fillna("").astype(str).str.contains(rf"(^|,\s*){re.escape(pcode)}(\s*,|$)", regex=True)
    item.loc[mask, "PACK당낱개수"] = str(int(forced_pack))

req_piece_item = (
    req_f.groupby(["년", "분기", "품목코드"], as_index=False)["요청수량_낱개환산"]
    .sum(min_count=1)
    .rename(columns={"요청수량_낱개환산": "요청수량_낱개"})
)
in_piece_item = (
    in_f.groupby(["년", "분기", "품목코드"], as_index=False)["출고수량_낱개환산"]
    .sum(min_count=1)
    .rename(columns={"출고수량_낱개환산": "출고수량_낱개"})
)
item = item.merge(req_piece_item, on=["년", "분기", "품목코드"], how="left")
item = item.merge(in_piece_item, on=["년", "분기", "품목코드"], how="left")
item_req_piece = pd.to_numeric(item["요청수량_낱개"], errors="coerce").fillna(0)
item_ship_piece = pd.to_numeric(item["출고수량_낱개"], errors="coerce").fillna(0)
item["매칭출고수량_낱개"] = np.minimum(item_req_piece, item_ship_piece)
item["초과출고수량_낱개"] = np.maximum(item_ship_piece - item_req_piece, 0)
item["잔량_낱개"] = np.maximum(item_req_piece - item_ship_piece, 0)

item = add_progress_columns(item, "요청수량_PACK", "출고수량_EA")
item = item.merge(item_product_master, on="품목코드", how="left")
item["제품코드(마스터)"] = item["제품코드(마스터)"].fillna("")

# 동일제품 통합(낱개 기준) 요약
# 집계 우선순위: 제품코드(마스터) -> (마스터코드가 비어있을 때만) P코드 -> 제품군명(최후 fallback)
family_src = item.copy()
family_src["제품코드_집계"] = family_src["제품코드(마스터)"].apply(normalize_key_value)
family_src["P코드_집계"] = family_src["P코드"].apply(normalize_key_value)
family_src["제품군명_집계"] = family_src["제품군명"].apply(normalize_key_value)
family_src["집계기준"] = np.where(
    family_src["제품코드_집계"] != "",
    "제품코드(마스터)",
    np.where(family_src["P코드_집계"] != "", "P코드", "제품군명"),
)
family_src["집계키"] = np.where(
    family_src["제품코드_집계"] != "",
    family_src["제품코드_집계"],
    np.where(family_src["P코드_집계"] != "", family_src["P코드_집계"], family_src["제품군명_집계"]),
)
family_src["집계키"] = family_src["집계키"].replace("", "(미분류)")
family_src["통합키"] = family_src["집계기준"] + ":" + family_src["집계키"]

family = (
    family_src.groupby(["년", "분기", "집계기준", "집계키", "통합키"], as_index=False)
    .agg(
        제품군명=("제품군명", summarize_names),
        대표품명=("품명", summarize_names),
        제품코드목록=("제품코드", summarize_codes),
        P코드=("P코드", summarize_codes),
        브랜드=("브랜드", summarize_codes),
        구분=("구분", summarize_codes),
        요청수량_PACK=("요청수량_PACK", "sum"),
        총출고수량_EA=("총출고수량_EA", "sum"),
        요청수량_낱개=("요청수량_낱개", "sum"),
        총출고수량_낱개=("출고수량_낱개", "sum"),
    )
)
family["제품군명"] = family["제품군명"].apply(normalize_key_value).replace("", "(미분류)")
family["대표품명"] = family["대표품명"].apply(normalize_key_value)
family["제품코드목록"] = family["제품코드목록"].apply(normalize_key_value)
family["P코드"] = family["P코드"].apply(normalize_key_value)
family["브랜드"] = family["브랜드"].apply(normalize_key_value)
family["구분"] = family["구분"].apply(normalize_key_value)

family = add_progress_columns(family, "요청수량_PACK", "총출고수량_EA")
family_req_piece = pd.to_numeric(family["요청수량_낱개"], errors="coerce").fillna(0)
family_ship_piece = pd.to_numeric(family["총출고수량_낱개"], errors="coerce").fillna(0)
family["매칭출고수량_낱개"] = np.minimum(family_req_piece, family_ship_piece)
family["초과출고수량_낱개"] = np.maximum(family_ship_piece - family_req_piece, 0)
family["잔량_낱개"] = np.maximum(family_req_piece - family_ship_piece, 0)
family["진행률_낱개(%)"] = np.where(family_req_piece > 0, (family["매칭출고수량_낱개"] / family_req_piece) * 100, np.nan)

global_search = str(st.session_state.get("global_search", ""))

# KPI (통합 검색 입력 시 검색결과 기준 요약)
kpi_source = item.copy()
has_global_terms = bool(parse_search_terms(global_search))
if has_global_terms:
    kpi_source = apply_or_search(
        kpi_source,
        global_search,
        ["제품코드", "품목코드", "제품코드(마스터)", "P코드", "브랜드", "구분", "품명", "품명검색", "상태", "PACK당낱개수", "년", "분기"],
    )

total_req = float(pd.to_numeric(kpi_source["요청수량_PACK"], errors="coerce").fillna(0).sum())
total_ship_total = float(pd.to_numeric(kpi_source["총출고수량_EA"], errors="coerce").fillna(0).sum())
total_ship_matched = float(pd.to_numeric(kpi_source["매칭출고수량_EA"], errors="coerce").fillna(0).sum())
total_ship_excess = float(pd.to_numeric(kpi_source["초과출고수량_EA"], errors="coerce").fillna(0).sum())
total_remaining = total_req - total_ship_matched
total_req_piece = float(pd.to_numeric(kpi_source["요청수량_낱개"], errors="coerce").fillna(0).sum())
total_ship_matched_piece = float(pd.to_numeric(kpi_source["매칭출고수량_낱개"], errors="coerce").fillna(0).sum())
total_remaining_piece = float(pd.to_numeric(kpi_source["잔량_낱개"], errors="coerce").fillna(0).sum())
progress_pct = (total_ship_matched_piece / total_req_piece * 100) if total_req_piece > 0 else 0.0

detail_mode_selected = str(st.session_state.get("tab2_detail_mode", "품목코드 상세"))
show_unified_inventory_kpi = detail_mode_selected == "동일제품 통합(낱개기준)"

total_stock_qty = 0.0
total_real_shortage = 0.0
if show_unified_inventory_kpi:
    inventory_stock_map_for_kpi, _ = load_inventory_stock(".")
    inventory_stock_dict_for_kpi = (
        inventory_stock_map_for_kpi.set_index("제품코드(마스터)")["보유재고"].to_dict()
        if not inventory_stock_map_for_kpi.empty
        else {}
    )
    family_kpi = family.copy()
    family_kpi = apply_or_search(
        family_kpi,
        global_search,
        ["집계기준", "집계키", "제품군명", "대표품명", "제품코드목록", "P코드", "브랜드", "구분", "상태", "년", "분기"],
    )
    family_kpi["보유재고"] = np.where(
        family_kpi["집계기준"] == "제품코드(마스터)",
        family_kpi["집계키"].apply(lambda v: calc_stock_sum_for_master_codes(v, inventory_stock_dict_for_kpi)),
        0,
    )
    family_kpi["재고반영수량"] = np.minimum(family_kpi["보유재고"], pd.to_numeric(family_kpi["잔량_낱개"], errors="coerce").fillna(0))
    family_kpi["실제부족량"] = np.maximum(pd.to_numeric(family_kpi["잔량_낱개"], errors="coerce").fillna(0) - family_kpi["재고반영수량"], 0)
    total_stock_qty = float(family_kpi["재고반영수량"].sum())
    total_real_shortage = float(family_kpi["실제부족량"].sum())

metric_cols = st.columns(10 if show_unified_inventory_kpi else 8)
metric_cols[0].metric("요청수량 (PACK)", f"{total_req:,.0f}")
metric_cols[1].metric("매칭출고수량 (PACK)", f"{total_ship_matched:,.0f}")
metric_cols[2].metric("초과출고수량 (PACK)", f"{total_ship_excess:,.0f}")
metric_cols[3].metric("총출고수량 (PACK)", f"{total_ship_total:,.0f}")
metric_cols[4].metric("잔량 (요청-매칭)", f"{total_remaining:,.0f}")
metric_cols[5].metric("출고율(매칭기준)", f"{progress_pct:,.1f}%")
render_color_metric(
    metric_cols[6],
    "요청수량 (낱개)",
    f"{total_req_piece:,.0f}",
    bg_color="#ecfeff",
    border_color="#67e8f9",
    label_color="#0e7490",
    value_color="#0f172a",
)
render_color_metric(
    metric_cols[7],
    "미출고 잔량(낱개)",
    f"{total_remaining_piece:,.0f}",
    bg_color="#fff1f2",
    border_color="#fda4af",
    label_color="#be123c",
    value_color="#7f1d1d",
)
if show_unified_inventory_kpi:
    render_color_metric(
        metric_cols[8],
        "보유재고(반영)",
        f"{total_stock_qty:,.0f}",
        bg_color="#ecfdf5",
        border_color="#86efac",
        label_color="#166534",
        value_color="#14532d",
    )
    render_color_metric(
        metric_cols[9],
        "실제부족량",
        f"{total_real_shortage:,.0f}",
        bg_color="#fff7ed",
        border_color="#fdba74",
        label_color="#9a3412",
        value_color="#7c2d12",
    )

item_count = len(kpi_source)
product_count = kpi_source["제품코드"].astype(str).replace("nan", "").replace("", np.nan).nunique(dropna=True)
status_counts = kpi_source["상태"].value_counts()
status_text = ", ".join([f"{k} {v}건" for k, v in status_counts.items()][:4]) if not status_counts.empty else "없음"
scope_text = "통합 검색 결과 요약" if has_global_terms else "전체 요약"
type_scope_text = f" | 구분필터: {selected_type}" if selected_type != "전체" else ""
st.caption(f"{scope_text}{type_scope_text} | 품목 {item_count:,}건 | 제품명코드 {product_count:,}개 | 상태분포: {status_text}")

global_search = st.text_input(
    "통합 검색 (OR)",
    value=global_search,
    key="global_search",
    placeholder="예: S036, Bandage, 미출고",
)
st.caption("쉼표(,) 또는 | 또는 / 로 키워드를 구분하면 OR 조건으로 검색합니다.")

# 탭 구성
tab1, tab2, tab3 = st.tabs(["제품명코드 요약", "품목코드 상세", "분기 누적 추이"])

with tab1:
    show_short_only = st.checkbox("미달(출고중/미출고)만 보기", value=False)
    prod_view = prod.copy()
    prod_view = apply_or_search(prod_view, global_search, ["제품코드", "P코드", "브랜드", "구분", "품명", "상태", "PACK당낱개수", "년", "분기"])
    if show_short_only:
        prod_view = prod_view[prod_view["상태"].isin(["미출고", "출고중"])]

    prod_view = prod_view.sort_values(["상태", "잔량"], ascending=[True, False])
    prod_cols = [
        "년",
        "분기",
        "제품코드",
        "P코드",
        "브랜드",
        "구분",
        "품명",
        "PACK당낱개수",
        "요청수량_PACK",
        "총출고수량_EA",
        "매칭출고수량_EA",
        "초과출고수량_EA",
        "잔량",
        "요청수량_낱개",
        "출고수량_낱개",
        "매칭출고수량_낱개",
        "초과출고수량_낱개",
        "잔량_낱개",
        "진행률(%)",
        "상태",
    ]
    st.dataframe(
        format_table(
            prod_view[prod_cols],
            int_cols=[
                "년",
                "분기",
                "요청수량_PACK",
                "총출고수량_EA",
                "매칭출고수량_EA",
                "초과출고수량_EA",
                "잔량",
                "요청수량_낱개",
                "출고수량_낱개",
                "매칭출고수량_낱개",
                "초과출고수량_낱개",
                "잔량_낱개",
            ],
            pct_cols=["진행률(%)"],
            progress_bar_cols=["진행률(%)"],
            status_col="상태",
            positive_alert_cols=["초과출고수량_EA", "초과출고수량_낱개", "잔량", "잔량_낱개"],
        ),
        use_container_width=True,
        hide_index=True,
        column_config={
            "상태": st.column_config.TextColumn("상태", width="large"),
            "진행률(%)": st.column_config.NumberColumn("출고율(%)", format="%.1f"),
            "제품코드": st.column_config.TextColumn("제품명코드"),
        },
    )
    excel_data_prod = to_excel_bytes(prod_view[prod_cols], sheet_name="제품코드요약")
    inventory_stock_map_for_pack_plan, _ = load_inventory_stock(".")
    inventory_stock_dict_for_pack_plan = (
        inventory_stock_map_for_pack_plan.set_index("제품코드(마스터)")["보유재고"].to_dict()
        if not inventory_stock_map_for_pack_plan.empty
        else {}
    )
    pcode_stock_totals_for_pack_plan = build_pcode_stock_totals(item, inventory_stock_dict_for_pack_plan)
    pcode_detail_map_for_pack_plan = build_pcode_detail_map(family)
    pack_plan_data = to_master_pack_plan_excel_bytes(
        prod,
        pcode_stock_totals_for_pack_plan,
        pcode_detail_map_for_pack_plan,
        req_f,
        item,
        inventory_stock_dict_for_pack_plan,
        in_f,
    )
    yymmdd = datetime.now().strftime("%y%m%d")
    dl_col1, dl_col2 = st.columns(2)
    with dl_col1:
        st.download_button(
            "제품명코드 요약 엑셀 다운로드",
            data=excel_data_prod,
            file_name="제품명코드_요약_요청대비출고.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with dl_col2:
        st.download_button(
            "포장계획 엑셀 다운로드",
            data=pack_plan_data,
            file_name=f"포장계획_{yymmdd}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    chart_all = prod_view.sort_values("잔량", ascending=False).copy()
    shortage_all = chart_all[chart_all["잔량"] > 0].copy()
    default_top_n = min(20, max(1, len(shortage_all))) if len(shortage_all) > 0 else 1

    c_left, c_right = st.columns([2, 1])
    with c_left:
        chart_mode = st.radio(
            "그래프 표시 범위",
            ["전체 제품", "상위 부족 제품"],
            horizontal=True,
            key="chart_mode",
        )
    with c_right:
        top_n = st.number_input(
            "상위 부족 제품 개수",
            min_value=1,
            max_value=max(1, len(shortage_all)),
            value=default_top_n,
            step=1,
            key="chart_top_n",
        )

    if chart_mode == "전체 제품":
        chart_source = chart_all
        chart_title = f"전체 제품 잔량 그래프 ({len(chart_source):,}개)"
    else:
        chart_source = shortage_all.head(int(top_n))
        chart_title = f"상위 부족 {int(top_n):,}개 제품"

    st.subheader(chart_title)
    if chart_source.empty:
        st.info("표시할 제품이 없습니다.")
    else:
        chart_df = chart_source[["제품코드", "품명", "잔량"]].copy()
        chart_df["대표품명"] = chart_df["품명"].astype(str).str.split(",").str[0].str.strip()
        chart_df["대표품명"] = chart_df["대표품명"].replace("", "(품명없음)")
        chart_df["그래프표시명"] = chart_df.apply(
            lambda r: f"{r['제품코드']} | {r['대표품명'][:24]}{'...' if len(r['대표품명']) > 24 else ''}",
            axis=1,
        )
        y_max = float(pd.to_numeric(chart_df["잔량"], errors="coerce").fillna(0).max())
        y_domain_max = max(1.0, y_max * 1.12)
        x_enc = alt.X(
            "그래프표시명:N",
            sort="-y",
            title="제품명코드 | 제품명",
            axis=alt.Axis(labelAngle=-90, labelPadding=8, labelLimit=220),
        )
        y_enc = alt.Y("잔량:Q", title="잔량", scale=alt.Scale(domain=[0, y_domain_max]))
        bar = alt.Chart(chart_df).mark_bar().encode(
            x=x_enc,
            y=y_enc,
            tooltip=[
                alt.Tooltip("제품코드:N", title="제품명코드"),
                alt.Tooltip("품명:N", title="품명"),
                alt.Tooltip("잔량:Q", title="잔량", format=","),
            ],
        )
        label = alt.Chart(chart_df).mark_text(
            dy=-10,
            font=APP_FONT_FAMILY,
            fontSize=14,
            fontWeight="bold",
            color="#111827",
        ).encode(
            x=x_enc,
            y=alt.Y("잔량:Q", scale=alt.Scale(domain=[0, y_domain_max])),
            text=alt.Text("잔량:Q", format=","),
        )
        # 상위 N개 모드에서 차트가 급격히 작아 보이지 않도록 최소 높이를 더 크게 유지
        chart_height = max(760, min(1400, 28 * len(chart_df)))
        bar_chart = apply_chart_style(
            (bar + label).properties(
                height=chart_height,
                padding={"top": 60, "bottom": 190, "left": 20, "right": 20},
            )
        )
        st.altair_chart(bar_chart, use_container_width=True)

with tab2:
    detail_mode = st.radio(
        "보기 방식",
        ["품목코드 상세", "동일제품 통합(낱개기준)", "세부사항", "생산요청 원본리스트", "입고수량 원본리스트"],
        horizontal=True,
        key="tab2_detail_mode",
    )

    if detail_mode == "품목코드 상세":
        item_view = item.copy()
        item_view = apply_or_search(
            item_view,
            global_search,
            ["제품코드", "품목코드", "제품코드(마스터)", "P코드", "브랜드", "구분", "품명", "품명검색", "제품군명", "상태", "PACK당낱개수", "년", "분기"],
        )

        item_view = item_view.sort_values(["상태", "잔량"], ascending=[True, False])
        item_cols = [
            "년",
            "분기",
            "제품코드",
            "P코드",
            "브랜드",
            "구분",
            "품목코드",
            "제품코드(마스터)",
            "품명",
            "PACK당낱개수",
            "요청수량_PACK",
            "총출고수량_EA",
            "매칭출고수량_EA",
            "초과출고수량_EA",
            "잔량",
            "요청수량_낱개",
            "출고수량_낱개",
            "매칭출고수량_낱개",
            "초과출고수량_낱개",
            "잔량_낱개",
            "진행률(%)",
            "상태",
        ]
        st.dataframe(
            format_table(
                item_view[item_cols],
                int_cols=[
                    "년",
                    "분기",
                    "요청수량_PACK",
                    "총출고수량_EA",
                    "매칭출고수량_EA",
                    "초과출고수량_EA",
                    "잔량",
                    "요청수량_낱개",
                    "출고수량_낱개",
                    "매칭출고수량_낱개",
                    "초과출고수량_낱개",
                    "잔량_낱개",
                ],
                pct_cols=["진행률(%)"],
                progress_bar_cols=["진행률(%)"],
                status_col="상태",
                positive_alert_cols=["초과출고수량_EA", "초과출고수량_낱개", "잔량", "잔량_낱개"],
            ),
            use_container_width=True,
            hide_index=True,
            column_config={
                "상태": st.column_config.TextColumn("상태", width="large"),
                "진행률(%)": st.column_config.NumberColumn("출고율(%)", format="%.1f"),
                "제품코드": st.column_config.TextColumn("제품명코드"),
                "제품코드(마스터)": st.column_config.TextColumn("제품코드"),
            },
        )

        excel_data = to_excel_bytes(item_view[item_cols], sheet_name="품목코드상세")
        st.download_button(
            "품목코드 상세 엑셀 다운로드",
            data=excel_data,
            file_name="품목코드_요청대비출고_상세.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif detail_mode == "동일제품 통합(낱개기준)":
        st.caption("집계 기준: 제품코드(마스터) 우선, 마스터코드가 없는 항목만 P코드 기준으로 통합합니다.")
        family_view = family.copy()
        inventory_stock_map, stock_file = load_inventory_stock(".")
        inventory_stock_dict = (
            inventory_stock_map.set_index("제품코드(마스터)")["보유재고"].to_dict()
            if not inventory_stock_map.empty
            else {}
        )
        family_view["보유재고"] = np.where(
            family_view["집계기준"] == "제품코드(마스터)",
            family_view["집계키"].apply(lambda v: calc_stock_sum_for_master_codes(v, inventory_stock_dict)),
            0,
        )
        family_view["재고반영수량"] = np.minimum(family_view["보유재고"], pd.to_numeric(family_view["잔량_낱개"], errors="coerce").fillna(0))
        family_view["실제부족량"] = np.maximum(pd.to_numeric(family_view["잔량_낱개"], errors="coerce").fillna(0) - family_view["재고반영수량"], 0)
        if stock_file:
            st.caption(f"재고 반영 파일: `{stock_file}`")
        family_view = apply_or_search(
            family_view,
            global_search,
            ["집계기준", "집계키", "제품군명", "대표품명", "제품코드목록", "P코드", "브랜드", "구분", "상태", "년", "분기"],
        )
        family_view = family_view.sort_values(["상태", "잔량_낱개"], ascending=[True, False])
        family_cols = [
            "년",
            "분기",
            "집계기준",
            "집계키",
            "제품군명",
            "대표품명",
            "제품코드목록",
            "P코드",
            "브랜드",
            "구분",
            "요청수량_낱개",
            "총출고수량_낱개",
            "매칭출고수량_낱개",
            "초과출고수량_낱개",
            "잔량_낱개",
            "보유재고",
            "실제부족량",
            "진행률_낱개(%)",
            "요청수량_PACK",
            "총출고수량_EA",
            "매칭출고수량_EA",
            "초과출고수량_EA",
            "상태",
        ]
        st.dataframe(
            format_table(
                family_view[family_cols],
                int_cols=[
                    "년",
                    "분기",
                    "요청수량_낱개",
                    "총출고수량_낱개",
                    "매칭출고수량_낱개",
                    "초과출고수량_낱개",
                    "잔량_낱개",
                    "보유재고",
                    "실제부족량",
                    "요청수량_PACK",
                    "총출고수량_EA",
                    "매칭출고수량_EA",
                    "초과출고수량_EA",
                ],
                pct_cols=["진행률_낱개(%)"],
                progress_bar_cols=["진행률_낱개(%)"],
                status_col="상태",
                positive_alert_cols=["초과출고수량_EA", "초과출고수량_낱개", "잔량_낱개", "실제부족량"],
            ),
            use_container_width=True,
            hide_index=True,
            column_config={
                "상태": st.column_config.TextColumn("상태", width="large"),
                "진행률_낱개(%)": st.column_config.NumberColumn("출고율_낱개(%)", format="%.1f"),
                "제품코드목록": st.column_config.TextColumn("제품명코드목록"),
            },
        )

        family_excel = to_excel_bytes(family_view[family_cols], sheet_name="동일제품통합")
        st.download_button(
            "동일제품 통합 엑셀 다운로드",
            data=family_excel,
            file_name="동일제품_통합_낱개기준.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif detail_mode == "세부사항":
        st.caption("세부사항: P코드 > S코드 > 개입/품명 > 세부코드(파워) 순으로 파워별 오더/재고를 표시합니다.")
        detail_view = item.copy()
        inventory_stock_map, stock_file = load_inventory_stock(".")
        inventory_stock_dict = (
            inventory_stock_map.set_index("제품코드(마스터)")["보유재고"].to_dict()
            if not inventory_stock_map.empty
            else {}
        )
        detail_view["세부코드"] = detail_view["제품코드(마스터)"].apply(normalize_key_value)
        detail_view["세부코드"] = np.where(detail_view["세부코드"] != "", detail_view["세부코드"], detail_view["품목코드"])
        detail_view["보유재고"] = detail_view["세부코드"].apply(lambda v: calc_stock_sum_for_master_codes(v, inventory_stock_dict))
        detail_view["재고반영수량"] = np.minimum(detail_view["보유재고"], pd.to_numeric(detail_view["잔량_낱개"], errors="coerce").fillna(0))
        detail_view["실제부족량"] = np.maximum(pd.to_numeric(detail_view["잔량_낱개"], errors="coerce").fillna(0) - detail_view["재고반영수량"], 0)
        detail_view["판매"] = detail_view["제품코드"]
        detail_view["생산"] = detail_view["P코드"].apply(normalize_key_value)
        detail_view["품명"] = detail_view["품명"].astype(str).str.split(",").str[0].str.strip()
        detail_view["개입"] = detail_view["PACK당낱개수"].apply(extract_pack_count).fillna(1).astype(int)
        if stock_file:
            st.caption(f"재고 반영 파일: `{stock_file}`")
        detail_view = apply_or_search(
            detail_view,
            global_search,
            ["세부코드", "판매", "생산", "개입", "품명", "품목코드", "제품코드(마스터)", "P코드", "브랜드", "구분", "상태", "년", "분기"],
        )
        detail_group_cols = ["판매", "생산", "개입", "품명", "세부코드"]
        detail_sum_cols = [
            "요청수량_PACK",
            "요청수량_낱개",
            "총출고수량_EA",
            "출고수량_낱개",
            "매칭출고수량_EA",
            "매칭출고수량_낱개",
            "초과출고수량_EA",
            "초과출고수량_낱개",
            "잔량_낱개",
            "보유재고",
            "실제부족량",
        ]
        detail_view = (
            detail_view.groupby(detail_group_cols, as_index=False)[detail_sum_cols]
            .sum(min_count=1)
            .fillna(0)
        )
        detail_view["_pcode_group_size"] = detail_view.groupby("생산", dropna=False)["생산"].transform("size")
        detail_view = detail_view.sort_values(
            ["_pcode_group_size", "생산", "판매", "개입", "품명", "세부코드"],
            ascending=[False, True, True, False, True, True],
        )
        detail_display = detail_view.rename(
            columns={
                "품명": "품 명",
                "요청수량_PACK": "총 오더(팩)",
                "요청수량_낱개": "총 오더(pcs)",
                "총출고수량_EA": "용마 총입고수량(팩)",
                "출고수량_낱개": "용마 총입고수량(pcs)",
                "잔량_낱개": "오더 잔량(pcs)",
                "보유재고": "낱개 재고(본사)",
                "실제부족량": "생산 필요량(pcs)",
            }
        )
        detail_cols = [
            "생산",
            "판매",
            "개입",
            "품 명",
            "세부코드",
            "총 오더(팩)",
            "총 오더(pcs)",
            "용마 총입고수량(팩)",
            "용마 총입고수량(pcs)",
            "오더 잔량(pcs)",
            "낱개 재고(본사)",
            "생산 필요량(pcs)",
        ]
        # 세부사항 화면은 실제 셀 병합(rowspan) 형태로 렌더링
        table_df = detail_display[detail_cols].copy()
        int_cols = [
            "개입",
            "총 오더(팩)",
            "총 오더(pcs)",
            "용마 총입고수량(팩)",
            "용마 총입고수량(pcs)",
            "오더 잔량(pcs)",
            "낱개 재고(본사)",
            "생산 필요량(pcs)",
        ]
        merge_cols = ["생산", "판매", "개입", "품 명"]
        n_rows = len(table_df)
        n_cols = len(detail_cols)
        col_idx = {c: i for i, c in enumerate(detail_cols)}
        rowspan = {(r, c): 1 for r in range(n_rows) for c in range(n_cols)}
        skip_cell = set()
        for m_i, m_col in enumerate(merge_cols):
            c = col_idx[m_col]
            prev_cols = merge_cols[:m_i]
            start = 0
            while start < n_rows:
                end = start + 1
                while end < n_rows:
                    same_curr = str(table_df.iloc[end][m_col]) == str(table_df.iloc[start][m_col])
                    same_prev = all(str(table_df.iloc[end][pc]) == str(table_df.iloc[start][pc]) for pc in prev_cols)
                    if not (same_curr and same_prev):
                        break
                    end += 1
                span = end - start
                if span > 1:
                    rowspan[(start, c)] = span
                    for r in range(start + 1, end):
                        skip_cell.add((r, c))
                start = end

        def fmt_val(col: str, val: object) -> str:
            if pd.isna(val):
                return ""
            if col in int_cols:
                try:
                    return f"{int(round(float(val))):,}"
                except Exception:
                    return str(val)
            return str(val)

        header_html = "".join(f"<th>{html.escape(c)}</th>" for c in detail_cols)
        body_rows = []
        for r in range(n_rows):
            tds = []
            for c in range(n_cols):
                if (r, c) in skip_cell:
                    continue
                col = detail_cols[c]
                val = fmt_val(col, table_df.iloc[r, c])
                rs = rowspan.get((r, c), 1)
                rs_attr = f' rowspan="{rs}"' if rs > 1 else ""
                tds.append(f"<td{rs_attr}>{html.escape(val)}</td>")
            body_rows.append("<tr>" + "".join(tds) + "</tr>")
        merged_table_html = f"""
        <div style="overflow-x:auto;">
          <table style="border-collapse:collapse; width:100%; font-size:14px;">
            <thead><tr>{header_html}</tr></thead>
            <tbody>{''.join(body_rows)}</tbody>
          </table>
        </div>
        <style>
          table th, table td {{ border:1px solid #cbd5e1; padding:6px 8px; white-space:nowrap; }}
          table th {{ background:#f8fafc; font-weight:700; text-align:center; }}
          table td {{ text-align:right; }}
          table td:nth-child(1), table td:nth-child(2), table td:nth-child(4), table td:nth-child(5) {{ text-align:left; }}
        </style>
        """
        st.markdown(merged_table_html, unsafe_allow_html=True)

        detail_excel = to_excel_bytes(
            detail_display[detail_cols],
            sheet_name="세부사항",
            merge_cols=["생산", "판매", "개입", "품 명"],
        )
        st.download_button(
            "세부사항 엑셀 다운로드",
            data=detail_excel,
            file_name="세부사항_낱개기준.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif detail_mode == "생산요청 원본리스트":
        st.caption("2026 국내 생산요청수량리스트 원본 데이터를 그대로 표시합니다.")
        raw_view = req_raw.copy()
        if not raw_view.empty:
            raw_cols = [c for c in raw_view.columns]
            raw_view = apply_or_search(raw_view, global_search, raw_cols)
            st.dataframe(raw_view, use_container_width=True, hide_index=True)
            raw_excel = to_excel_bytes(raw_view, sheet_name="생산요청원본")
            st.download_button(
                "생산요청 원본리스트 엑셀 다운로드",
                data=raw_excel,
                file_name="생산요청_원본리스트.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("생산요청 원본 데이터가 없습니다.")
    else:
        st.caption("2026년 국내제품 입고수량 엑셀 원본 데이터를 그대로 표시합니다.")
        inbound_view = inbound_raw.copy()
        if not inbound_view.empty:
            inbound_cols = [c for c in inbound_view.columns]
            inbound_view = apply_or_search(inbound_view, global_search, inbound_cols)
            st.dataframe(inbound_view, use_container_width=True, hide_index=True)
            inbound_excel = to_excel_bytes(inbound_view, sheet_name="입고수량원본")
            st.download_button(
                "입고수량 원본리스트 엑셀 다운로드",
                data=inbound_excel,
                file_name="입고수량_원본리스트.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("입고수량 원본 데이터가 없습니다.")

with tab3:
    trend_options = ["일별", "주별", "월별"]
    if "trend_granularity" not in st.session_state:
        st.session_state["trend_granularity"] = "일별"
    for option in trend_options:
        state_key = f"trend_granularity_{option}"
        if state_key not in st.session_state:
            st.session_state[state_key] = (option == st.session_state["trend_granularity"])

    def _set_trend_granularity(selected_option: str) -> None:
        st.session_state["trend_granularity"] = selected_option
        for opt in trend_options:
            st.session_state[f"trend_granularity_{opt}"] = (opt == selected_option)

    selector_wrap_col, _ = st.columns([1.8, 8.2])
    with selector_wrap_col:
        g_col1, g_col2, g_col3 = st.columns(3, gap="small")
        with g_col1:
            st.checkbox("일별", key="trend_granularity_일별", on_change=_set_trend_granularity, args=("일별",))
        with g_col2:
            st.checkbox("주별", key="trend_granularity_주별", on_change=_set_trend_granularity, args=("주별",))
        with g_col3:
            st.checkbox("월별", key="trend_granularity_월별", on_change=_set_trend_granularity, args=("월별",))

    selected_granularity = st.session_state.get("trend_granularity", "일별")

    trend_base = in_f.dropna(subset=["이동일자"]).copy()
    if selected_granularity == "주별":
        trend_base["집계일"] = trend_base["이동일자"].dt.to_period("W-SUN").dt.start_time
        period_col_name = "주차"
        x_axis_title = "주차(월요일 시작)"
    elif selected_granularity == "월별":
        trend_base["집계일"] = trend_base["이동일자"].dt.to_period("M").dt.to_timestamp()
        period_col_name = "월"
        x_axis_title = "월"
    else:
        trend_base["집계일"] = trend_base["이동일자"].dt.normalize()
        period_col_name = "일자"
        x_axis_title = "이동일자"

    cumulative = trend_base.groupby("집계일", as_index=True)["출고수량_EA"].sum().sort_index().cumsum()
    trend_export_df = pd.DataFrame(columns=[period_col_name, "누적출고(EA)", "요청수량(PACK)"])
    if cumulative.empty:
        st.info("표시할 이동일자 데이터가 없습니다.")
    else:
        trend_df = pd.DataFrame({"누적출고(EA)": cumulative})
        if total_req > 0:
            trend_df["요청수량(PACK)"] = total_req
        trend_reset = trend_df.reset_index().rename(columns={"집계일": period_col_name})
        trend_export_df = trend_reset.copy()
        trend_long = trend_reset.melt(id_vars=period_col_name, var_name="구분", value_name="수량")

        line = alt.Chart(trend_long).mark_line(point=True).encode(
            x=alt.X(f"{period_col_name}:T", title=x_axis_title),
            y=alt.Y("수량:Q", title="수량"),
            color=alt.Color("구분:N", title="지표"),
            tooltip=[
                alt.Tooltip(f"{period_col_name}:T", title=period_col_name),
                alt.Tooltip("구분:N", title="지표"),
                alt.Tooltip("수량:Q", title="수량", format=","),
            ],
        )
        last_points = trend_long.sort_values(period_col_name).groupby("구분", as_index=False).tail(1)
        labels = alt.Chart(last_points).mark_text(
            dx=6,
            align="left",
            font=APP_FONT_FAMILY,
            fontSize=14,
            fontWeight="bold",
        ).encode(
            x=alt.X(f"{period_col_name}:T"),
            y=alt.Y("수량:Q"),
            color=alt.Color("구분:N", legend=None),
            text=alt.Text("수량:Q", format=","),
        )
        line_chart = apply_chart_style((line + labels).properties(height=360))
        st.altair_chart(line_chart, use_container_width=True)

    excel_data_trend = to_excel_bytes(trend_export_df, sheet_name="분기누적추이")
    st.download_button(
        "분기 누적 추이 엑셀 다운로드",
        data=excel_data_trend,
        file_name="분기_누적출고_추이.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.caption("입고 원본은 사용자 요청에 따라 중복 제거 없이 그대로 집계합니다.")
